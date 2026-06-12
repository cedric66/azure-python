"""Offline tests for the sandbox command family (clone, impact, k8s-test, spot-sim,
upgrade-rehearsal, kubectl helpers). No Azure, no az/kubectl/kubelogin binaries.

  uv run python tests/test_sandbox.py
"""
import argparse
import base64
import json
import os
import re
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import smoke_test as st  # reuse the 3-cluster fleet fixture + requests stub

import azrep.http_client as hc
import azrep.kubectl as kctl
from azrep import sandbox, sandbox_clone, sandbox_impact, sandbox_spot, sandbox_upgrade

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMP = tempfile.mkdtemp(prefix="akssbx_")
OUT = os.path.join(TMP, "reports")


def expect(cond, msg):
    assert cond, msg


# ------------------------------------------------------------------ fakes
class FakeSession:
    def __init__(self, handler):
        self.handler = handler
        self.calls = []

    def request(self, method, url, *, params=None, payload=None, ok404=False,
                min_interval=None):
        self.calls.append((method, url, params, payload))
        return self.handler(method, url, params, payload)

    def get(self, url, **kw):
        return self.request("GET", url, **kw)

    def post(self, url, **kw):
        return self.request("POST", url, **kw)

    def put(self, url, **kw):
        return self.request("PUT", url, **kw)

    def delete(self, url, **kw):
        return self.request("DELETE", url, **kw)

    def get_paged(self, url, params=None, value_key="value"):
        return (self.get(url, params=params) or {}).get(value_key) or []

    def post_paged(self, url, payload, value_key="value"):
        return (self.post(url, payload=payload) or {}).get(value_key) or []

    def puts(self, fragment):
        return [c for c in self.calls if c[0] == "PUT" and fragment in c[1].lower()]


class FakeRun:
    """Scripted subprocess.run replacement for azrep.kubectl._run."""

    def __init__(self, script):
        self.script = list(script)  # [(rc, stdout, stderr), ...]
        self.argvs = []
        self.envs = []

    def __call__(self, argv, capture_output=True, text=True, timeout=None, env=None):
        self.argvs.append(list(argv))
        self.envs.append(env or {})
        rc, out, err = self.script.pop(0) if self.script else (0, "", "")
        return argparse.Namespace(returncode=rc, stdout=out, stderr=err)


def make_cfg(**over):
    cfg = {
        "subscription_id": st.S1,
        "resource_group": "rg-aks-sandbox-dev",
        "location": "eastus",
        "cluster": {"name": "aks-sbx-01",
                    "node_pools": [{"name": "sys", "mode": "System",
                                    "vm_size": "Standard_D4s_v5", "count": 1}]},
        "_config_path": os.path.join(TMP, "sandbox.yaml"),
        "_config_dir": REPO,
    }
    cfg.update(over)
    return cfg


def ns(**kw):
    return argparse.Namespace(**kw)


# ------------------------------------------------------------------ build_node_pool
def test_build_node_pool_taints_labels():
    out = sandbox.build_node_pool({
        "name": "usr", "vm_size": "Standard_D4s_v5", "count": 2,
        "node_taints": ["team=payments:NoSchedule"],
        "node_labels": {"team": "payments"},
    })
    expect(out["nodeTaints"] == ["team=payments:NoSchedule"], "nodeTaints missing: %s" % out)
    expect(out["nodeLabels"] == {"team": "payments"}, "nodeLabels missing: %s" % out)
    print("OK build_node_pool taints/labels")


# ------------------------------------------------------------------ clone
def clone_handler(method, url, params, payload):
    if "microsoft.resourcegraph" in url.lower():
        m = re.search(r"tolower\(id\) == '([^']+)'", payload["query"])
        rows = [c for c in st.CLUSTERS if c["id"].lower() == m.group(1)]
        return {"data": rows, "$skipToken": None}
    raise AssertionError("clone should be ARG-only, got %s %s" % (method, url))


def run_clone(cluster_id, **over):
    import yaml
    args = ns(cluster_id=cluster_id, base=None, name=None,
              out=os.path.join(TMP, "clone.yaml"), keep_subnets=False,
              keep_sku_tier=False, keep_counts=False)
    for k, v in over.items():
        setattr(args, k, v)
    hc.connect = lambda min_interval=0.05: FakeSession(clone_handler)
    sandbox_clone.run(args)
    with open(args.out, encoding="utf-8") as f:
        return yaml.safe_load(f)


def test_clone_prod_cluster():
    cfg = run_clone(st.CL3)
    cl = cfg["cluster"]
    expect(cl["name"].startswith("sbx-aks-prod-01"), "clone name: %s" % cl["name"])
    expect(cfg["subscription_id"].startswith("TODO"), "no base -> TODO subscription")
    expect(cl["sku_tier"] == "Free", "tier must downsize to Free")
    expect(cl["kubernetes_version"] == "1.32.1", "version must be preserved")
    expect(cl["authorized_ip_ranges"] == [], "authorized ranges must be stripped")
    expect(cl["network"]["plugin"] == "azure" and cl["network"]["policy"] == "azure",
           "CNI shape must be preserved: %s" % cl["network"])
    expect(cl["aad_profile"] == {"managed": True, "enableAzureRBAC": False},
           "aad profile mapping: %s" % cl.get("aad_profile"))
    pools = {p["name"]: p for p in cl["node_pools"]}
    expect(set(pools) == {"sys", "app", "bat"}, "pools: %s" % set(pools))
    expect(all(p["count"] == 1 for p in pools.values()), "counts must downsize to 1")
    expect(pools["app"]["min_count"] == 0 and pools["app"]["max_count"] == 2,
           "user autoscaler must downsize to 0..2: %s" % pools["app"])
    expect("min_count" not in pools["sys"], "non-autoscaled system pool has no min/max")
    expect(pools["bat"]["priority"] == "Spot" and pools["bat"]["eviction_policy"] == "Delete",
           "spot pool must be preserved: %s" % pools["bat"])
    expect("node_taints" not in pools["bat"], "auto spot taint must be dropped")
    expect(all("vnet_subnet_id" not in p for p in pools.values()),
           "subnet IDs must be stripped by default")
    expect(cfg["tags"]["cloned_from"] == st.CL3, "cloned_from tag")
    print("OK clone prod cluster")


def test_clone_private_cilium_udr():
    cfg = run_clone(st.CL2, name="sbx-clone-02")
    cl = cfg["cluster"]
    expect(cl["private_cluster"] is True, "private cluster kept")
    expect(cl["network"]["dataplane"] == "cilium", "dataplane kept (with warning)")
    expect(cl["network"]["outbound_type"] == "loadBalancer",
           "UDR must downgrade to loadBalancer when subnets stripped: %s" % cl["network"])
    expect(cl["disable_local_accounts"] is True, "local accounts setting kept")
    print("OK clone private/cilium/UDR cluster")


def test_clone_keep_flags():
    cfg = run_clone(st.CL3, keep_counts=True, keep_subnets=True, keep_sku_tier=True)
    pools = {p["name"]: p for p in cfg["cluster"]["node_pools"]}
    expect(pools["app"]["count"] == 6, "--keep-counts must keep source counts")
    expect(pools["app"]["vnet_subnet_id"] == st.S2_PROD_SUBNET,
           "--keep-subnets must keep subnet IDs")
    expect(cfg["cluster"]["sku_tier"] == "Standard", "--keep-sku-tier must keep tier")
    print("OK clone keep flags")


def test_clone_rejects_non_sandbox_name():
    try:
        run_clone(st.CL3, name="aks-production-clone")
    except SystemExit as e:
        expect("sandbox" in str(e), "should explain the sandbox-name rule: %s" % e)
        print("OK clone rejects non-sandbox name")
        return
    raise AssertionError("clone accepted a non-sandbox name")


# ------------------------------------------------------------------ impact
def raw_rows():
    rows = []
    for c in st.CLUSTERS:
        rows.append({"id": c["id"], "name": c["name"],
                     "type": "Microsoft.ContainerService/managedClusters",
                     "location": c["location"], "tags": c["tags"],
                     "sku": {"name": c["skuName"], "tier": c["skuTier"]},
                     "identity": {"type": c["identityType"]},
                     "properties": {"apiServerAccessProfile": c["apiServerAccessProfile"]},
                     "subscriptionId": c["subscriptionId"]})
    return rows


def impact_handler(assignment_id):
    def handler(method, url, params, payload):
        low = url.lower()
        if "microsoft.resourcegraph" in low:
            q = payload["query"]
            if "project id, name, type" in q:
                data = raw_rows()
            elif "managedclusters" in q:
                data = st.CLUSTERS
            elif "subscriptions/resourcegroups" in q:
                data = st.RGS
            else:
                data = st.SUBNAMES
            subs = [s.lower() for s in payload["subscriptions"]]
            return {"data": [d for d in data if d["subscriptionId"].lower() in subs],
                    "$skipToken": None}
        if "checkpolicyrestrictions" in low:
            content = payload["resourceDetails"]["resourceContent"]
            private = (content["properties"].get("apiServerAccessProfile") or {}) \
                .get("enablePrivateCluster")
            evals = [] if private else [
                {"policyInfo": {"policyAssignmentId": assignment_id},
                 "evaluationResult": "NonCompliant",
                 "evaluationDetails": {"evaluatedExpressions": [
                     {"path": "properties.apiServerAccessProfile.enablePrivateCluster",
                      "operator": "Equals", "expressionValue": False}]}}]
            return {"fieldRestrictions": [],
                    "contentEvaluationResult": {"policyEvaluations": evals}}
        if method in ("PUT", "DELETE") and "policy" in low:
            return {}
        raise AssertionError("unmocked impact call: %s %s" % (method, url))
    return handler


def test_impact():
    cfg = make_cfg()
    assignment_id = "%s/providers/Microsoft.Authorization/policyAssignments/impact-candidate" \
        % sandbox.rg_scope(cfg)
    csv_path = os.path.join(TMP, "subscriptions.csv")
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("subscription_id,subscription_name,include\n")
        f.write("%s,contoso-platform,Y\n%s,contoso-shared,Y\n" % (st.S1, st.S2))
    session = FakeSession(impact_handler(assignment_id))
    args = ns(policy=os.path.join(REPO, "policies", "audit-aks-private-api.json"),
              name="impact-candidate", params=["effect=Audit"], effect_override=False,
              keep_assignment=False, csv=csv_path, all=True, subs=None, env=None,
              nonprod=False, out=OUT)
    path = sandbox_impact.run(session, cfg, args)

    checks = [c for c in session.calls if "checkpolicyrestrictions" in c[1].lower()]
    expect(len(checks) == 3, "3 clusters -> 3 evaluations, got %d" % len(checks))
    for _m, _u, params, payload in checks:
        expect(payload["includeAuditEffect"] is True, "includeAuditEffect must be true")
        expect(payload["resourceDetails"]["apiVersion"] == sandbox_impact.AKS_BODY_API,
               "resourceDetails.apiVersion missing")
        expect(params["api-version"] == sandbox_impact.CHECK_API, "check api-version")
    def_puts = session.puts("policydefinitions/impact-candidate")
    asn_puts = session.puts("policyassignments/impact-candidate")
    expect(def_puts and asn_puts, "candidate definition+assignment must be staged")
    expect(asn_puts[0][3]["properties"]["enforcementMode"] == "DoNotEnforce",
           "staged assignment must be DoNotEnforce")
    deletes = [c for c in session.calls if c[0] == "DELETE"]
    expect(len(deletes) == 2, "teardown must delete assignment+definition: %s" % deletes)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["PerCluster"]
    hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
    rcol = hdr.index("result") + 1
    res = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=rcol).value
           for r in range(2, ws.max_row + 1)}
    expect(res == {"aks-dev-01": "NonCompliant", "aks-prod-01": "NonCompliant",
                   "aks-dev-02": "Compliant"},
           "impact results wrong: %s" % res)
    print("OK impact sweep")


def test_impact_effect_override():
    payload = {"properties": {"parameters": {"effect": {
        "type": "String", "allowedValues": ["Audit", "Disabled"], "defaultValue": "Audit"}}}}
    out = sandbox_impact.override_effect(payload)
    eff = out["properties"]["parameters"]["effect"]
    expect(eff["defaultValue"] == "Deny" and "Deny" in eff["allowedValues"],
           "override must force Deny default: %s" % eff)
    payload2 = {"properties": {"parameters": {"effect": {"allowedValues": ["audit", "deny"]}}}}
    eff2 = sandbox_impact.override_effect(payload2)["properties"]["parameters"]["effect"]
    expect(eff2["defaultValue"] == "deny", "lowercase deny variants must be reused: %s" % eff2)
    print("OK impact effect override")


# ------------------------------------------------------------------ k8s-test
def test_k8s_cases():
    from azrep import sandbox_k8s
    cfg = make_cfg()
    case_deny = {"name": "deny", "manifest": "policies/tests/pod-bad-registry.yaml",
                 "expect": "deny"}
    case_allow = {"name": "allow", "manifest": "policies/tests/pod-good-registry.yaml",
                  "expect": "allow"}
    old_have, old_run = kctl.have, kctl._run
    kctl.have = lambda b: True
    try:
        # deny + webhook rejection -> PASS
        kctl._run = FakeRun([(1, "", 'admission webhook "validation.gatekeeper.sh" '
                                     "denied the request: bad registry")])
        name, _e, status, detail = sandbox_k8s.run_case(
            "/kc", cfg, case_deny, "policy-test", 1)
        expect(status == "PASS", "deny case should PASS on webhook rejection: %s" % detail)

        # deny but admitted -> FAIL (and a cleanup delete is issued)
        fake = FakeRun([(0, "pod/x created", ""), (0, "", "")])
        kctl._run = fake
        _n, _e, status, detail = sandbox_k8s.run_case("/kc", cfg, case_deny, "policy-test", 1)
        expect(status == "FAIL" and "admitted" in detail, "deny case admitted: %s" % detail)
        expect(any("delete" in a for a in fake.argvs[1]), "cleanup delete expected")

        # allow + admitted -> PASS
        kctl._run = FakeRun([(0, "pod/x created", ""), (0, "", "")])
        _n, _e, status, detail = sandbox_k8s.run_case("/kc", cfg, case_allow, "policy-test", 1)
        expect(status == "PASS", "allow case should PASS: %s" % detail)

        # allow but denied -> FAIL
        kctl._run = FakeRun([(1, "", "denied the request"), (0, "", "")])
        _n, _e, status, detail = sandbox_k8s.run_case("/kc", cfg, case_allow, "policy-test", 1)
        expect(status == "FAIL", "allow case denied should FAIL: %s" % detail)

        # audit: admitted, violation shows up in constraints
        constraints = {"items": [{
            "kind": "K8sAzureV2ContainerAllowedImages",
            "metadata": {"name": "azurepolicy-k8sazurev2containerallowedimages-x"},
            "spec": {"enforcementAction": "dryrun"},
            "status": {"violations": [{"kind": "Pod", "name": "policy-test-bad-registry",
                                       "namespace": "policy-test",
                                       "message": "image not allowed"}]},
        }]}
        case_audit = dict(case_deny, name="audit", expect="audit")
        kctl._run = FakeRun([(0, "pod/x created", ""),
                             (0, json.dumps(constraints), ""), (0, "", "")])
        _n, _e, status, detail = sandbox_k8s.run_case("/kc", cfg, case_audit, "policy-test", 1)
        expect(status == "PASS" and "image not allowed" in detail,
               "audit case should find the violation: %s" % detail)

        # KUBECONFIG isolation on every kubectl call
        kctl._run = fake = FakeRun([(0, "{}", "")])
        kctl.run_kubectl("/tmp/kc-test", ["get", "nodes"])
        expect(fake.envs[0].get("KUBECONFIG") == "/tmp/kc-test",
               "KUBECONFIG must point at the dedicated file")
    finally:
        kctl.have, kctl._run = old_have, old_run
    print("OK k8s-test cases")


def test_kubeconfig_arm_fallback():
    kubeconfig = "apiVersion: v1\nkind: Config\nclusters: []\n"

    def handler(method, url, params, payload):
        if "listclusterusercredential" in url.lower():
            return {"kubeconfigs": [{"name": "user", "value":
                    base64.b64encode(kubeconfig.encode()).decode()}]}
        raise AssertionError("unexpected call %s" % url)

    cfg = make_cfg(_config_dir=TMP)
    old_have = kctl.have
    kctl.have = lambda b: False  # no az, no kubelogin
    try:
        path = kctl.fetch_kubeconfig(cfg, session=FakeSession(handler), force=True)
        with open(path, encoding="utf-8") as f:
            expect(f.read() == kubeconfig, "kubeconfig must be the decoded ARM payload")
        expect(os.path.basename(path) == ".kubeconfig-aks-sbx-01", "dedicated file name")
    finally:
        kctl.have = old_have
    print("OK kubeconfig ARM fallback")


# ------------------------------------------------------------------ spot-sim
def spot_handler(state):
    wrk = {"name": "wrk", "properties": {
        "mode": "User", "vmSize": "Standard_D4s_v3", "count": 3,
        "enableAutoScaling": True, "minCount": 1, "maxCount": 5,
        "availabilityZones": ["1", "2"], "osType": "Linux", "osSKU": "Ubuntu",
        "maxPods": 110, "orchestratorVersion": "1.31.5",
        "powerState": {"code": "Running"}, "provisioningState": "Succeeded"}}
    sys_pool = {"name": "sys", "properties": {"mode": "System", "vmSize": "Standard_D4s_v5",
                "count": 1, "powerState": {"code": "Running"},
                "provisioningState": "Succeeded"}}

    def handler(method, url, params, payload):
        low = url.lower()
        if method == "GET" and low.endswith("/agentpools"):
            return {"value": [sys_pool, wrk]}
        if method == "PUT" and "/agentpools/" in low:
            state[url.rsplit("/", 1)[1]] = payload
            return payload
        if method == "GET" and "/agentpools/" in low:
            name = url.rsplit("/", 1)[1].split("?")[0]
            if name == "wrk":
                return wrk
            put = state.get(name)
            return {"name": name, "properties": dict((put or {}).get("properties", {}),
                                                     provisioningState="Succeeded")}
        raise AssertionError("unmocked spot call: %s %s" % (method, url))
    return handler


def test_spot_sim_payloads():
    state = {}
    session = FakeSession(spot_handler(state))
    cfg = make_cfg()
    args = ns(pool="wrk", vm_size=None, spot_share=0.6, scenarios="none",
              simulate_eviction=False, keep=False, md=False, no_prices=True, out=OUT)
    sandbox_spot.run(session, cfg, args)

    spot_puts = [c for c in session.calls
                 if c[0] == "PUT" and "/agentpools/" in c[1].lower()
                 and not c[1].lower().endswith("/wrk")]
    expect(len(spot_puts) == 1, "exactly one new spot pool PUT, got %d" % len(spot_puts))
    props = spot_puts[0][3]["properties"]
    expect(props["scaleSetPriority"] == "Spot" and props["scaleSetEvictionPolicy"] == "Delete"
           and props["spotMaxPrice"] == -1, "spot pool PUT body: %s" % props)
    expect(props["mode"] == "User" and props["enableAutoScaling"] is True, "spot pool mode")
    # od_floor 0.4 of 3 nodes -> keep 2 OD, 1 initial spot, max ceil((5-2)*1.3)=4
    expect(props["count"] == 1 and props["minCount"] == 0 and props["maxCount"] == 4,
           "spot sizing from plan_split: %s" % props)
    expect(props["nodeLabels"].get("pooltype") == "spot", "pooltype label expected")
    expect(props["orchestratorVersion"] == "1.31.5", "spot pool matches OD pool version")
    expect("nodeTaints" not in props, "AKS adds the spot taint itself")

    od_puts = [c for c in session.calls
               if c[0] == "PUT" and c[1].lower().endswith("/agentpools/wrk")]
    expect(len(od_puts) == 1, "OD shrink PUT expected")
    od = od_puts[0][3]["properties"]
    expect(od["minCount"] == 1 and od["maxCount"] == 2,
           "OD autoscaler shrink to 1..2 (keep 2): %s" % od)
    print("OK spot-sim ARM payloads")


def fake_plan():
    return {"team": "sandbox", "od_pool": "wrk", "spot_pool": "sandboxsp",
            "vm_size": "Standard_D4s_v3", "zones": "1,2", "current_nodes": 3,
            "current_max": 5, "od_keep_nodes": 2, "spot_initial_nodes": 1,
            "spot_min": 0, "spot_max": 4, "spot_share_target": "33%",
            "team_taints": [], "labels": {}}


def test_spot_scenarios():
    import yaml
    plan = fake_plan()
    ids = [s["id"] for s in sandbox_spot.SCENARIOS]
    expect(len(ids) == len(set(ids)) == 10, "10 unique scenarios: %s" % ids)
    tolerated = {"toleration-only", "spot-preferred", "spot-required", "topology-spread",
                 "spread-soft", "single-replica-spot", "pdb-too-strict"}
    spread = {"topology-spread", "spread-soft", "spread-missing-toleration",
              "pdb-too-strict"}
    for sc in sandbox_spot.SCENARIOS:
        docs = sc["build"](plan, "spot-sim-%s" % sc["id"])
        yaml.safe_load_all(yaml.safe_dump_all(docs))  # round-trips as valid YAML
        dep = docs[0]
        spec = dep["spec"]["template"]["spec"]
        has_tol = any(t.get("value") == "spot" for t in spec.get("tolerations", []))
        expect(has_tol == (sc["id"] in tolerated),
               "%s toleration presence wrong (%s)" % (sc["id"], has_tol))
        has_spread = bool(spec.get("topologySpreadConstraints"))
        expect(has_spread == (sc["id"] in spread),
               "%s spread presence wrong" % sc["id"])
        if has_spread:
            tsc = spec["topologySpreadConstraints"][0]
            expect(tsc["topologyKey"] == sandbox_spot.POOL_LABEL,
                   "%s must spread on the agentpool label, not scalesetpriority" % sc["id"])
        if sc["id"] == "pdb-too-strict":
            expect(docs[1]["kind"] == "PodDisruptionBudget"
                   and docs[1]["spec"]["minAvailable"] == "100%", "PDB doc expected")

    def pods(*specs):
        return [{"pod": "p%d" % i, "node": n, "pool": pl, "spot": s, "phase": ph,
                 "message": msg}
                for i, (n, pl, s, ph, msg) in enumerate(specs)]

    ok, detail = sandbox_spot.v_skew(pods(("n1", "wrk", False, "Running", ""),
                                          ("n2", "sandboxsp", True, "Running", "")), plan)
    expect(ok == "PASS", "balanced spread should PASS: %s" % detail)
    bad, detail = sandbox_spot.v_skew(pods(("n1", "wrk", False, "Running", ""),
                                           ("n2", "wrk", False, "Running", ""),
                                           ("n3", "wrk", False, "Running", "")), plan)
    expect(bad == "FAIL", "skew 3-0 should FAIL: %s" % detail)
    ok, detail = sandbox_spot.v_expect_pending(
        pods(("n1", "wrk", False, "Running", ""),
             ("", "", False, "Pending", "0/3 nodes available: untolerated taint")), plan)
    expect(ok == "PASS", "missing-toleration Pending should be detected: %s" % detail)
    ok, detail = sandbox_spot.v_all_running(
        pods(("n1", "wrk", False, "Running", ""), ("n2", "wrk", False, "Running", "")),
        expect_spot=False)
    expect(ok == "PASS", "od-pinned 0-on-spot should PASS: %s" % detail)
    warn, detail = sandbox_spot.v_single_replica(
        pods(("n1", "sandboxsp", True, "Running", "")), plan)
    expect(warn == "WARN", "single replica on spot should WARN: %s" % detail)
    print("OK spot scenario matrix")


# ------------------------------------------------------------------ upgrade
SUPPORTED = {
    "1.30": {"patches": ["1.30.9"], "support_plans": [], "is_preview": False,
             "is_default": False},
    "1.31": {"patches": ["1.31.5", "1.31.6"], "support_plans": [], "is_preview": False,
             "is_default": False},
    "1.32": {"patches": ["1.32.1", "1.32.2"], "support_plans": [], "is_preview": False,
             "is_default": True},
    "1.33": {"patches": ["1.33.0"], "support_plans": [], "is_preview": True,
             "is_default": False},
}


def test_hop_path():
    expect(sandbox_upgrade.hop_path("1.29.4", "1.32", SUPPORTED)
           == ["1.30.9", "1.31.6", "1.32.2"], "3 hops with highest patches")
    expect(sandbox_upgrade.hop_path("1.29.4", "next", SUPPORTED) == ["1.30.9"], "next = +1")
    expect(sandbox_upgrade.hop_path("1.31.5", "latest", SUPPORTED) == ["1.32.2"],
           "latest skips previews")
    expect(sandbox_upgrade.hop_path("1.30.9", "1.32.1", SUPPORTED)
           == ["1.31.6", "1.32.1"], "explicit target patch honored")
    for cur, tgt in (("1.32.1", "1.32"), ("1.32.2", "next")):
        try:
            sandbox_upgrade.hop_path(cur, tgt, SUPPORTED)
            raise AssertionError("expected SystemExit for %s -> %s" % (cur, tgt))
        except SystemExit:
            pass
    gap = {k: v for k, v in SUPPORTED.items() if k != "1.31"}
    try:
        sandbox_upgrade.hop_path("1.29.4", "1.32", gap)
        raise AssertionError("missing intermediate minor must error")
    except SystemExit as e:
        expect("intermediate" in str(e), "gap error message: %s" % e)
    print("OK upgrade hop path")


def test_control_plane_put_strips_pools():
    state = {}

    def handler(method, url, params, payload):
        if method == "GET":
            return {"location": "eastus", "sku": {"name": "Base", "tier": "Free"},
                    "identity": {"type": "SystemAssigned"}, "tags": {},
                    "properties": {"kubernetesVersion": "1.29.4",
                                   "provisioningState": "Succeeded",
                                   "currentKubernetesVersion": "1.29.4",
                                   "fqdn": "x.azmk8s.io", "maxAgentPools": 100,
                                   "agentPoolProfiles": [{"name": "sys"}]}}
        if method == "PUT":
            state["put"] = payload
            return payload
        raise AssertionError("unmocked %s %s" % (method, url))

    sandbox_upgrade.upgrade_control_plane(FakeSession(handler), make_cfg(), "1.30.9")
    props = state["put"]["properties"]
    expect("agentPoolProfiles" not in props, "control-plane PUT must omit agentPoolProfiles")
    expect(props["kubernetesVersion"] == "1.30.9", "target version set")
    expect("provisioningState" not in props and "fqdn" not in props,
           "read-only props must be stripped")
    print("OK control-plane-only PUT")


def test_deprecated_api_check():
    path = os.path.join(TMP, "old.yaml")
    with open(path, "w", encoding="utf-8") as f:
        f.write("apiVersion: policy/v1beta1\nkind: PodSecurityPolicy\n"
                "metadata: {name: psp}\n---\n"
                "apiVersion: flowcontrol.apiserver.k8s.io/v1beta3\nkind: FlowSchema\n"
                "metadata: {name: fs}\n---\n"
                "apiVersion: apps/v1\nkind: Deployment\nmetadata: {name: ok}\n")
    found = sandbox_upgrade.deprecated_api_check([path], ["1.25.0", "1.32.2"])
    kinds = {(f[1], f[2]) for f in found}
    expect(("policy/v1beta1", "PodSecurityPolicy") in kinds,
           "PSP removal must be flagged: %s" % kinds)
    expect(("flowcontrol.apiserver.k8s.io/v1beta3", "FlowSchema") in kinds,
           "flowcontrol v1beta3 removal must be flagged: %s" % kinds)
    expect(len(found) == 2, "the apps/v1 Deployment is fine: %s" % found)
    expect(sandbox_upgrade.deprecated_api_check([path], ["1.30.9"]) == [],
           "no findings when no crossed minor removes them")
    print("OK deprecated API check")


def main():
    os.makedirs(OUT, exist_ok=True)
    test_build_node_pool_taints_labels()
    test_clone_prod_cluster()
    test_clone_private_cilium_udr()
    test_clone_keep_flags()
    test_clone_rejects_non_sandbox_name()
    test_impact()
    test_impact_effect_override()
    test_k8s_cases()
    test_kubeconfig_arm_fallback()
    test_spot_sim_payloads()
    test_spot_scenarios()
    test_hop_path()
    test_control_plane_put_strips_pools()
    test_deprecated_api_check()
    print("\nALL SANDBOX TESTS PASSED")


if __name__ == "__main__":
    main()
