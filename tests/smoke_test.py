"""Offline smoke test: runs every report script end-to-end against mocked Azure
API responses and validates the generated workbooks. No Azure access needed.

  uv run python tests/smoke_test.py
"""
import datetime as dt
import json
import os
import re
import sys
import tempfile
import types

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import requests  # noqa: F401
except ModuleNotFoundError:
    requests_stub = types.ModuleType("requests")
    requests_stub.RequestException = Exception
    requests_stub.Session = lambda: None
    sys.modules["requests"] = requests_stub

import azrep.http_client as hc
import azrep.costmgmt as costmgmt
import azrep.armextras as armextras

S1 = "11111111-1111-1111-1111-111111111111"
S2 = "22222222-2222-2222-2222-222222222222"
TODAY = dt.date.today()
MONTHS = []
y, m = TODAY.year, TODAY.month - 3
while m <= 0:
    m += 12
    y -= 1
for _ in range(4):
    MONTHS.append("%04d-%02d" % (y, m))
    m += 1
    if m > 12:
        m, y = 1, y + 1

CL1 = ("/subscriptions/%s/resourcegroups/rg-apps-dev/providers/"
       "microsoft.containerservice/managedclusters/aks-dev-01" % S1)
CL2 = ("/subscriptions/%s/resourcegroups/rg-data-dev/providers/"
       "microsoft.containerservice/managedclusters/aks-dev-02" % S1)
CL3 = ("/subscriptions/%s/resourcegroups/rg-prod/providers/"
       "microsoft.containerservice/managedclusters/aks-prod-01" % S2)
S1_DEV_SUBNET = ("/subscriptions/%s/resourceGroups/rg-network/providers/"
                 "Microsoft.Network/virtualNetworks/vnet-dev/subnets/aks-dev-nodes" % S1)
S1_DATA_SUBNET = ("/subscriptions/%s/resourceGroups/rg-network/providers/"
                  "Microsoft.Network/virtualNetworks/vnet-data/subnets/aks-data-nodes" % S1)
S2_PROD_SUBNET = ("/subscriptions/%s/resourceGroups/rg-network-prod/providers/"
                  "Microsoft.Network/virtualNetworks/vnet-prod/subnets/aks-prod-nodes" % S2)


def pool(name, mode, size, count, priority="Regular", zones=(), autoscale=False,
         image="AKSUbuntu-2204gen2containerd-202601.07.0", power="Running", **kw):
    p = {"name": name, "mode": mode, "vmSize": size, "count": count,
         "scaleSetPriority": priority, "availabilityZones": list(zones),
         "enableAutoScaling": autoscale, "osType": "Linux", "osSKU": "Ubuntu",
         "osDiskType": "Managed", "osDiskSizeGB": 128, "maxPods": kw.get("max_pods", 110),
         "orchestratorVersion": kw.get("ver", "1.31.5"),
         "currentOrchestratorVersion": kw.get("ver", "1.31.5"),
         "nodeImageVersion": image, "type": "VirtualMachineScaleSets",
         "powerState": {"code": power}}
    if priority == "Spot":
        p["spotMaxPrice"] = -1
        p["scaleSetEvictionPolicy"] = "Delete"
        p["nodeTaints"] = ["kubernetes.azure.com/scalesetpriority=spot:NoSchedule"]
        p["nodeLabels"] = {"kubernetes.azure.com/scalesetpriority": "spot"}
    if autoscale:
        p["minCount"], p["maxCount"] = 1, max(count, 5)
    if kw.get("vnet"):
        p["vnetSubnetID"] = kw["vnet"]
    if kw.get("pod_subnet"):
        p["podSubnetID"] = kw["pod_subnet"]
    return p


CLUSTERS = [
    {"id": CL1, "name": "aks-dev-01", "resourceGroup": "rg-apps-dev",
     "subscriptionId": S1, "location": "eastus", "tags": {"environment": "dev"},
     "kubernetesVersion": "1.29.4", "currentKubernetesVersion": "1.29.4",
     "provisioningState": "Succeeded", "powerState": "Running",
     "skuName": "Base", "skuTier": "Free", "supportPlan": "KubernetesOfficial",
     "nodeResourceGroup": "MC_rg-apps-dev_aks-dev-01_eastus",
     "dnsPrefix": "aksdev01", "fqdn": "aksdev01.hcp.eastus.azmk8s.io",
     "privateFQDN": "", "enableRBAC": True, "disableLocalAccounts": False,
     "identityType": "SystemAssigned", "servicePrincipalClientId": "",
     "agentPoolProfiles": [pool("sys", "System", "Standard_D4s_v3", 2, ver="1.29.4",
                                vnet=S1_DEV_SUBNET),
                           pool("wrk", "User", "Standard_D4s_v3", 3, zones=("1", "2"),
                                autoscale=True, ver="1.29.4", vnet=S1_DEV_SUBNET),
                           pool("spt", "User", "Standard_D4as_v4", 2, priority="Spot",
                                ver="1.29.4", vnet=S1_DEV_SUBNET)],
     "networkProfile": {"networkPlugin": "kubenet", "loadBalancerSku": "standard",
                        "outboundType": "loadBalancer"},
     "apiServerAccessProfile": {}, "aadProfile": {},
     "addonProfiles": {"azurepolicy": {"enabled": False},
                       "omsagent": {"enabled": False}},
     "autoUpgradeProfile": {}, "securityProfile": {},
     "oidcIssuerProfile": {"enabled": False},
     "autoScalerProfile": {"expander": "priority",
                           "balance-similar-node-groups": "true",
                           "scan-interval": "10s",
                           "scale-down-unneeded-time": "10m"}},
    {"id": CL2, "name": "aks-dev-02", "resourceGroup": "rg-data-dev",
     "subscriptionId": S1, "location": "eastus2", "tags": {},
     "kubernetesVersion": "1.31.5", "currentKubernetesVersion": "1.31.5",
     "provisioningState": "Succeeded", "powerState": "Stopped",
     "skuName": "Base", "skuTier": "Standard", "supportPlan": "KubernetesOfficial",
     "nodeResourceGroup": "MC_rg-data-dev_aks-dev-02_eastus2",
     "dnsPrefix": "aksdev02", "fqdn": "", "privateFQDN": "aksdev02.private.azmk8s.io",
     "enableRBAC": True, "disableLocalAccounts": True,
     "identityType": "UserAssigned", "servicePrincipalClientId": "",
     "agentPoolProfiles": [pool("sys", "System", "Standard_D4s_v5", 3,
                                zones=("1", "2", "3"), power="Stopped",
                                vnet=S1_DATA_SUBNET),
                           pool("usr", "User", "Standard_D8s_v5", 4,
                                zones=("1", "2", "3"), autoscale=True, power="Stopped",
                                vnet=S1_DATA_SUBNET)],
     "networkProfile": {"networkPlugin": "azure", "networkPluginMode": "overlay",
                        "networkPolicy": "cilium", "networkDataplane": "cilium",
                        "loadBalancerSku": "standard", "outboundType": "userDefinedRouting"},
     "apiServerAccessProfile": {"enablePrivateCluster": True},
     "aadProfile": {"managed": True, "enableAzureRBAC": True},
     "addonProfiles": {"azurepolicy": {"enabled": True}, "omsagent": {"enabled": True},
                       "azureKeyvaultSecretsProvider": {"enabled": True}},
     "autoUpgradeProfile": {"upgradeChannel": "patch", "nodeOSUpgradeChannel": "NodeImage"},
     "securityProfile": {"workloadIdentity": {"enabled": True},
                         "defender": {"securityMonitoring": {"enabled": True}},
                         "imageCleaner": {"enabled": True}},
     "oidcIssuerProfile": {"enabled": True},
     "autoScalerProfile": {"expander": "least-waste",
                           "balance-similar-node-groups": "false",
                           "scan-interval": "10s"}},
    {"id": CL3, "name": "aks-prod-01", "resourceGroup": "rg-prod",
     "subscriptionId": S2, "location": "westeurope", "tags": {},
     "kubernetesVersion": "1.32.1", "currentKubernetesVersion": "1.32.1",
     "provisioningState": "Succeeded", "powerState": "Running",
     "skuName": "Base", "skuTier": "Standard", "supportPlan": "KubernetesOfficial",
     "nodeResourceGroup": "MC_rg-prod_aks-prod-01_westeurope",
     "dnsPrefix": "aksprod01", "fqdn": "aksprod01.hcp.westeurope.azmk8s.io",
     "privateFQDN": "", "enableRBAC": True, "disableLocalAccounts": True,
     "identityType": "SystemAssigned", "servicePrincipalClientId": "",
     "agentPoolProfiles": [pool("sys", "System", "Standard_D8s_v5", 3,
                                zones=("1", "2", "3"), ver="1.32.1",
                                vnet=S2_PROD_SUBNET),
                           pool("app", "User", "Standard_D16s_v5", 6,
                                zones=("1", "2", "3"), autoscale=True, ver="1.32.1",
                                vnet=S2_PROD_SUBNET),
                           pool("bat", "User", "Standard_D8s_v5", 2, priority="Spot",
                                ver="1.32.1", vnet=S2_PROD_SUBNET)],
     "networkProfile": {"networkPlugin": "azure", "networkPolicy": "azure",
                        "loadBalancerSku": "standard", "outboundType": "loadBalancer"},
     "apiServerAccessProfile": {"authorizedIPRanges": ["1.2.3.0/24", "4.5.6.0/24"]},
     "aadProfile": {"managed": True, "enableAzureRBAC": False},
     "addonProfiles": {"azurepolicy": {"enabled": False}, "omsagent": {"enabled": True}},
     "autoUpgradeProfile": {"upgradeChannel": "stable", "nodeOSUpgradeChannel": "NodeImage"},
     "securityProfile": {}, "oidcIssuerProfile": {"enabled": True},
     "autoScalerProfile": {"expander": "random",
                           "balance-similar-node-groups": "false",
                           "scale-down-utilization-threshold": "0.5"}},
]

RGS = [
    {"subscriptionId": S1, "name": "rg-apps-dev", "tags": {}},
    {"subscriptionId": S1, "name": "rg-data-dev", "tags": {"env": "dev"}},
    {"subscriptionId": S2, "name": "rg-prod", "tags": {"environment": "prod"}},
]
SUBNAMES = [
    {"subscriptionId": S1, "name": "contoso-platform"},
    {"subscriptionId": S2, "name": "contoso-shared"},
]
SUBNETS = [
    {"id": S1_DEV_SUBNET, "name": "aks-dev-nodes", "subscriptionId": S1,
     "resourceGroup": "rg-network", "location": "eastus", "vnet": "vnet-dev",
     "addressPrefix": "10.10.1.0/28", "addressPrefixes": [],
     "nsgId": "/subscriptions/%s/resourceGroups/rg-network/providers/Microsoft.Network/networkSecurityGroups/nsg-dev" % S1,
     "routeTableId": "", "natGatewayId": "",
     "privateEndpointNetworkPolicies": "Enabled", "serviceEndpoints": [], "delegations": []},
    {"id": S1_DATA_SUBNET, "name": "aks-data-nodes", "subscriptionId": S1,
     "resourceGroup": "rg-network", "location": "eastus2", "vnet": "vnet-data",
     "addressPrefix": "10.20.1.0/26", "addressPrefixes": [],
     "nsgId": "", "routeTableId": "", "natGatewayId": "",
     "privateEndpointNetworkPolicies": "Enabled", "serviceEndpoints": [], "delegations": []},
    {"id": S2_PROD_SUBNET, "name": "aks-prod-nodes", "subscriptionId": S2,
     "resourceGroup": "rg-network-prod", "location": "westeurope", "vnet": "vnet-prod",
     "addressPrefix": "10.30.1.0/24", "addressPrefixes": [],
     "nsgId": "/subscriptions/%s/resourceGroups/rg-network-prod/providers/Microsoft.Network/networkSecurityGroups/nsg-prod" % S2,
     "routeTableId": "/subscriptions/%s/resourceGroups/rg-network-prod/providers/Microsoft.Network/routeTables/rt-prod" % S2,
     "natGatewayId": "", "privateEndpointNetworkPolicies": "Enabled",
     "serviceEndpoints": [{"service": "Microsoft.Storage"}], "delegations": []},
]


def _resource(rid, name, typ, sub, rg, loc, sku_name="", sku_tier="", tags=None):
    return {"id": rid, "name": name, "type": typ.lower(), "subscriptionId": sub,
            "resourceGroup": rg, "location": loc, "kind": "", "sku_name": sku_name,
            "sku_tier": sku_tier, "provisioning_state": "Succeeded",
            "tags": tags or {}}


def _mk_resources():
    rows = []
    for c in CLUSTERS:
        rows.append(_resource(c["id"], c["name"], "microsoft.containerservice/managedclusters",
                              c["subscriptionId"], c["resourceGroup"], c["location"],
                              c["skuName"], c["skuTier"], c.get("tags")))
        nrg = c["nodeResourceGroup"]
        sid, loc = c["subscriptionId"], c["location"]
        rows.extend([
            _resource("/subscriptions/%s/resourceGroups/%s/providers/Microsoft.Network/loadBalancers/kubernetes" % (sid, nrg),
                      "kubernetes", "microsoft.network/loadbalancers", sid, nrg, loc, "standard"),
            _resource("/subscriptions/%s/resourceGroups/%s/providers/Microsoft.Network/publicIPAddresses/kubernetes-a1" % (sid, nrg),
                      "kubernetes-a1", "microsoft.network/publicipaddresses", sid, nrg, loc, "standard"),
            _resource("/subscriptions/%s/resourceGroups/%s/providers/Microsoft.Compute/disks/aks-osdisk-a1" % (sid, nrg),
                      "aks-osdisk-a1", "microsoft.compute/disks", sid, nrg, loc, "P30"),
        ])
        for p in c["agentPoolProfiles"]:
            rows.append(_resource(
                "/subscriptions/%s/resourceGroups/%s/providers/Microsoft.Compute/virtualMachineScaleSets/aks-%s-12345678-vmss"
                % (sid, nrg, p["name"]),
                "aks-%s-12345678-vmss" % p["name"],
                "microsoft.compute/virtualmachinescalesets", sid, nrg, loc, p["vmSize"]))
    rows.extend([
        _resource("/subscriptions/%s/resourceGroups/rg-network/providers/Microsoft.Network/virtualNetworks/vnet-dev" % S1,
                  "vnet-dev", "microsoft.network/virtualnetworks", S1, "rg-network", "eastus"),
        _resource("/subscriptions/%s/resourceGroups/rg-network/providers/Microsoft.Network/virtualNetworks/vnet-data" % S1,
                  "vnet-data", "microsoft.network/virtualnetworks", S1, "rg-network", "eastus2"),
        _resource("/subscriptions/%s/resourceGroups/rg-network-prod/providers/Microsoft.Network/virtualNetworks/vnet-prod" % S2,
                  "vnet-prod", "microsoft.network/virtualnetworks", S2, "rg-network-prod", "westeurope"),
    ])
    return rows


RESOURCES = _mk_resources()


def _mk_cost_records():
    """(rg_lower, pricing_model, meter, resource_id, month, usd)"""
    recs = []

    def add(rg, pm, meter, rid, vals):
        for mo, v in zip(MONTHS, vals):
            if v:
                recs.append({"rg": rg, "pm": pm, "meter": meter, "rid": rid,
                             "month": mo, "usd": float(v)})

    rg1 = "mc_rg-apps-dev_aks-dev-01_eastus"
    vm1 = "/subscriptions/%s/resourcegroups/%s/providers/microsoft.compute" % (S1, rg1)
    add(rg1, "OnDemand", "D4s v3", vm1 + "/virtualmachinescalesets/aks-wrk-11111111-vmss",
        [200, 205, 210, 70])
    add(rg1, "OnDemand", "D2s v3", vm1 + "/virtualmachinescalesets/aks-sys-11111111-vmss",
        [100, 0, 0, 0])
    add(rg1, "OnDemand", "D8s v5", vm1 + "/virtualmachinescalesets/aks-sys-22222222-vmss",
        [0, 0, 110, 35])
    add(rg1, "Spot", "D4as v4 Spot", vm1 + "/virtualmachinescalesets/aks-spt-11111111-vmss",
        [40, 42, 44, 15])
    add(rg1, "Reservation", "D4s v3", vm1 + "/virtualmachinescalesets/aks-wrk-11111111-vmss",
        [60, 60, 60, 20])
    add(rg1, "OnDemand", "Standard HDD Managed Disks",
        vm1 + "/disks/agentdisks", [25, 26, 25, 8])

    rg2 = "mc_rg-data-dev_aks-dev-02_eastus2"
    vm2 = "/subscriptions/%s/resourcegroups/%s/providers/microsoft.compute" % (S1, rg2)
    add(rg2, "OnDemand", "D4s v5", vm2 + "/virtualmachinescalesets/aks-sys-33333333-vmss",
        [100, 100, 100, 30])

    rg3 = "mc_rg-prod_aks-prod-01_westeurope"
    vm3 = "/subscriptions/%s/resourcegroups/%s/providers/microsoft.compute" % (S2, rg3)
    add(rg3, "OnDemand", "D16s v5", vm3 + "/virtualmachinescalesets/aks-app-44444444-vmss",
        [800, 820, 1300, 400])
    add(rg3, "Reservation", "D8s v5", vm3 + "/virtualmachinescalesets/aks-sys-44444444-vmss",
        [200, 200, 200, 66])
    add(rg3, "Spot", "D8s v5 Spot", vm3 + "/virtualmachinescalesets/aks-bat-44444444-vmss",
        [30, 31, 33, 11])

    # managed-cluster resource fee rows (live in the cluster's own RG)
    add("rg-apps-dev", "OnDemand", "Standard Uptime SLA", CL1, [0, 0, 0, 0])
    add("rg-data-dev", "OnDemand", "Standard Uptime SLA", CL2, [73, 73, 73, 24])
    add("rg-prod", "OnDemand", "Standard Uptime SLA", CL3, [73, 73, 73, 24])
    return recs


COST_RECORDS = _mk_cost_records()


# --- single-subscription re-architecture fixture (subscription_rearch) -------
# A small synthetic estate in S1 used only by the `rearch` report. Routed in the
# fake ARG handler by query markers so it never disturbs the AKS fixtures above.
RE_RG_APP = "rg-rearch-app-dev"
RE_RG_NET = "rg-rearch-net"
RE_LOC = "eastus"


def _rid(provider, name, rg=RE_RG_APP):
    return ("/subscriptions/%s/resourceGroups/%s/providers/%s/%s"
            % (S1, rg, provider, name))


RE_VM_RUNNING = _rid("Microsoft.Compute/virtualMachines", "vm-dev-app-01")
RE_VM_STOPPED = _rid("Microsoft.Compute/virtualMachines", "vm-dev-legacy-01")
RE_DISK_ORPHAN = _rid("Microsoft.Compute/disks", "disk-orphan-01")
RE_PIP_IDLE = _rid("Microsoft.Network/publicIPAddresses", "pip-idle-01", RE_RG_NET)
RE_ASP_EMPTY = _rid("Microsoft.Web/serverfarms", "asp-empty-01")
RE_SA_GRS = _rid("Microsoft.Storage/storageAccounts", "stgrsdev01")
RE_FW = _rid("Microsoft.Network/azureFirewalls", "fw-dev-01", RE_RG_NET)

RE_RESOURCES = [
    {"id": RE_VM_RUNNING, "name": "vm-dev-app-01",
     "type": "microsoft.compute/virtualmachines", "kind": "",
     "resourceGroup": RE_RG_APP, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "", "sku_tier": "", "sku_capacity": "", "tags": {"env": "dev"}},
    {"id": RE_VM_STOPPED, "name": "vm-dev-legacy-01",
     "type": "microsoft.compute/virtualmachines", "kind": "",
     "resourceGroup": RE_RG_APP, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "", "sku_tier": "", "sku_capacity": "", "tags": {"env": "dev"}},
    {"id": RE_DISK_ORPHAN, "name": "disk-orphan-01",
     "type": "microsoft.compute/disks", "kind": "",
     "resourceGroup": RE_RG_APP, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "Premium_LRS", "sku_tier": "", "sku_capacity": "", "tags": {}},
    {"id": RE_PIP_IDLE, "name": "pip-idle-01",
     "type": "microsoft.network/publicipaddresses", "kind": "",
     "resourceGroup": RE_RG_NET, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "Standard", "sku_tier": "", "sku_capacity": "", "tags": {}},
    {"id": RE_ASP_EMPTY, "name": "asp-empty-01",
     "type": "microsoft.web/serverfarms", "kind": "app",
     "resourceGroup": RE_RG_APP, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "P1v2", "sku_tier": "PremiumV2", "sku_capacity": "1", "tags": {}},
    {"id": RE_SA_GRS, "name": "stgrsdev01",
     "type": "microsoft.storage/storageaccounts", "kind": "StorageV2",
     "resourceGroup": RE_RG_APP, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "Standard_GRS", "sku_tier": "Standard", "sku_capacity": "",
     "tags": {"env": "dev"}},
    {"id": RE_FW, "name": "fw-dev-01",
     "type": "microsoft.network/azurefirewalls", "kind": "",
     "resourceGroup": RE_RG_NET, "location": RE_LOC, "subscriptionId": S1,
     "sku_name": "AZFW_VNet", "sku_tier": "Standard", "sku_capacity": "", "tags": {}},
]

RE_DISKS = [{"id": RE_DISK_ORPHAN, "name": "disk-orphan-01",
             "resourceGroup": RE_RG_APP, "location": RE_LOC,
             "sku_name": "Premium_LRS", "diskSizeGB": 256, "diskState": "Unattached",
             "timeCreated": "2025-01-10T00:00:00Z"}]
RE_PIPS = [{"id": RE_PIP_IDLE, "name": "pip-idle-01", "resourceGroup": RE_RG_NET,
            "location": RE_LOC, "sku_name": "Standard", "allocationMethod": "Static"}]
RE_NICS = []
RE_LBS = []
RE_VMS = [
    {"id": RE_VM_RUNNING, "name": "vm-dev-app-01", "resourceGroup": RE_RG_APP,
     "location": RE_LOC, "vmSize": "Standard_D4s_v5",
     "powerState": "PowerState/running", "tags": {"env": "dev"}},
    {"id": RE_VM_STOPPED, "name": "vm-dev-legacy-01", "resourceGroup": RE_RG_APP,
     "location": RE_LOC, "vmSize": "Standard_D2s_v5",
     "powerState": "PowerState/stopped", "tags": {"env": "dev"}},
]
RE_VMSS = []
RE_SNAPSHOTS = []
RE_ASPS = [{"id": RE_ASP_EMPTY, "name": "asp-empty-01", "resourceGroup": RE_RG_APP,
            "location": RE_LOC, "sku_name": "P1v2", "sku_tier": "PremiumV2",
            "sku_capacity": "1", "numberOfSites": 0}]
RE_GATEWAYS = [{"id": RE_FW, "name": "fw-dev-01",
                "type": "microsoft.network/azurefirewalls",
                "resourceGroup": RE_RG_NET, "location": RE_LOC,
                "sku_name": "AZFW_VNet", "sku_tier": "Standard", "gwType": ""}]
RE_STORAGE = [{"id": RE_SA_GRS, "name": "stgrsdev01", "resourceGroup": RE_RG_APP,
               "location": RE_LOC, "sku_name": "Standard_GRS", "sku_tier": "Standard",
               "kind": "StorageV2", "accessTier": "Hot"}]
RE_SQL = []
RE_ADVISOR = [{"id": "/subscriptions/%s/providers/Microsoft.Advisor/recommendations/adv-1" % S1,
               "category": "Cost", "impact": "High",
               "problem": "Right-size or shut down underutilized virtual machines",
               "solution": "Resize vm-dev-app-01 to a smaller SKU",
               "impactedResource": RE_VM_RUNNING,
               "impactedType": "Microsoft.Compute/virtualMachines",
               "annualSavings": "1200", "savingsCurrency": "USD"}]
RE_RGS = [{"subscriptionId": S1, "name": RE_RG_APP.lower(), "tags": {"env": "dev"}},
          {"subscriptionId": S1, "name": RE_RG_NET.lower(), "tags": {}}]

# Per-resource cost rows for the rearch fixture: (resource_id, monthly USD).
RE_COST = {
    RE_VM_RUNNING: 300.0, RE_VM_STOPPED: 80.0, RE_DISK_ORPHAN: 35.0,
    RE_PIP_IDLE: 4.0, RE_ASP_EMPTY: 120.0, RE_SA_GRS: 60.0, RE_FW: 900.0,
}
RE_SERVICES = {
    RE_VM_RUNNING: "Virtual Machines", RE_VM_STOPPED: "Virtual Machines",
    RE_DISK_ORPHAN: "Storage", RE_PIP_IDLE: "Virtual Network",
    RE_ASP_EMPTY: "Azure App Service", RE_SA_GRS: "Storage",
    RE_FW: "Azure Firewall",
}


# (query marker -> rearch fixture). Order matters: the ALL_RESOURCES projection
# is matched last because it shares `tolower(type)` with several others.
_REARCH_ROUTES = [
    ("advisorresources", lambda: RE_ADVISOR),
    ("diskstate", lambda: RE_DISKS),
    ("ipconfiguration", lambda: RE_PIPS),
    ("isempty(properties.privateendpoint)", lambda: RE_NICS),
    ("backendaddresspools", lambda: RE_LBS),
    ("extend powerstate", lambda: RE_VMS),
    ("snapshots", lambda: RE_SNAPSHOTS),
    ("serverfarms", lambda: RE_ASPS),
    ("azurefirewalls", lambda: RE_GATEWAYS),
    ("storageaccounts", lambda: RE_STORAGE),
    ("servers/databases", lambda: RE_SQL),
]


def _rearch_kind(q):
    """Return a rearch route key for q, or None if it is not a rearch query."""
    ql = q.lower()
    # VMSS query: scale sets projected with sku_capacity but no extend/where body.
    if ("virtualmachinescalesets" in ql and "sku_capacity = tostring(sku.capacity)" in ql
            and "where type =~ 'microsoft.compute/virtualmachinescalesets'" in ql):
        return "vmss"
    for marker, _ in _REARCH_ROUTES:
        if marker in ql:
            return marker
    # ALL_RESOURCES_KQL: bare project with sku_capacity, no where clause.
    if ("sku_capacity = tostring(sku.capacity)" in ql and "tolower(type)" in ql
            and " where " not in ql.replace("\n", " ")):
        return "all"
    return None


def _is_rearch(q):
    return _rearch_kind(q) is not None


def _rearch_arg(q, subs):
    """Route a subscription_rearch ARG query to its fixture (S1 only)."""
    kind = _rearch_kind(q)
    if S1 not in subs:
        return []
    if kind == "all":
        return RE_RESOURCES
    if kind == "vmss":
        return RE_VMSS
    for marker, getter in _REARCH_ROUTES:
        if kind == marker:
            return getter()
    return []


def _rearch_cost(payload, groups):
    """subscription_rearch cost: ResourceId/ServiceName grouping over RE_COST.
    Each resource's monthly cost lands in every full month; the current MTD
    month gets a smaller partial so MoM trend has a real previous month."""
    gran = payload["dataset"].get("granularity", "Monthly")
    cur = TODAY.strftime("%Y-%m")
    agg = {}
    for rid, monthly in RE_COST.items():
        svc = RE_SERVICES.get(rid, "Other")
        for mo in MONTHS:
            v = monthly * (0.3 if mo == cur else 1.0)
            key = tuple({"ResourceId": rid, "ServiceName": svc}[g] for g in groups)
            period = "%s-01T00:00:00" % mo
            agg[(key, period)] = agg.get((key, period), 0.0) + v
    columns = ([{"name": "Cost", "type": "Number"}, {"name": "CostUSD", "type": "Number"}]
               + [{"name": g, "type": "String"} for g in groups]
               + [{"name": "BillingMonth", "type": "String"},
                  {"name": "Currency", "type": "String"}])
    rows = [[round(v, 4), round(v, 4)] + list(k) + [period, "USD"]
            for (k, period), v in sorted(agg.items(), key=lambda kv: str(kv[0]))]
    return {"properties": {"columns": columns, "rows": rows, "nextLink": None}}


def _cost_response(url, payload):
    scope = url.split("/providers/Microsoft.CostManagement")[0].lower()
    sub = scope.split("/subscriptions/")[1].split("/")[0]
    rg_scope = scope.split("/resourcegroups/")[1] if "/resourcegroups/" in scope else None
    ds = payload["dataset"]
    gran = ds.get("granularity", "Monthly")
    groups = [g["name"] for g in ds.get("grouping", [])]
    metric = payload.get("type", "AmortizedCost")
    if "ServiceName" in groups:
        return _rearch_cost(payload, groups)

    recs = [r for r in COST_RECORDS if r["rid"].lower().startswith("/subscriptions/" + sub)]
    if rg_scope:
        recs = [r for r in recs if r["rg"] == rg_scope]
    if metric == "ActualCost":
        recs = [r for r in recs if r["pm"] != "Reservation"]

    flt = ds.get("filter")
    if flt:
        dims = flt.get("dimensions") or {}
        name, values = dims.get("name", ""), [v.lower() for v in dims.get("values", [])]
        if name == "ResourceGroupName":
            recs = [r for r in recs if r["rg"] in values]
        elif name == "ResourceId":
            recs = [r for r in recs if r["rid"].lower() in values]
        elif name == "ResourceType":
            recs = [r for r in recs if "/managedclusters/" in r["rid"].lower()]

    def keyfor(r):
        out = []
        for g in groups:
            out.append({"ResourceGroupName": r["rg"], "PricingModel": r["pm"],
                        "Meter": r["meter"], "ResourceId": r["rid"]}[g])
        return tuple(out)

    agg = {}
    for r in recs:
        if gran == "Daily":
            for day, frac in (("05", 0.6), ("15", 0.4)):
                period = int("%s%s" % (r["month"].replace("-", ""), day))
                k = keyfor(r) + (period,)
                agg[k] = agg.get(k, 0.0) + r["usd"] * frac
        else:
            period = "%s-01T00:00:00" % r["month"]
            k = keyfor(r) + (period,)
            agg[k] = agg.get(k, 0.0) + r["usd"]

    datecol = "UsageDate" if gran == "Daily" else "BillingMonth"
    columns = ([{"name": "Cost", "type": "Number"}, {"name": "CostUSD", "type": "Number"}]
               + [{"name": g, "type": "String"} for g in groups]
               + [{"name": datecol, "type": "Number" if gran == "Daily" else "String"},
                  {"name": "Currency", "type": "String"}])
    rows = [[round(v, 4), round(v, 4)] + list(k[:-1]) + [k[-1], "USD"]
            for k, v in sorted(agg.items(), key=lambda kv: str(kv[0]))]
    return {"properties": {"columns": columns, "rows": rows, "nextLink": None}}


def _metrics_response(url, params):
    rid = url.split("/providers/microsoft.insights/metrics")[0].lower()
    stats = {CL1.lower(): (3.0, 12.0), CL3.lower(): (45.0, 62.0)}.get(rid, (20.0, 35.0))
    cpu, mem = stats
    names = params["metricnames"].split(",")
    base = dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=3)
    out = []
    for nm in names:
        val = {"node_cpu_usage_percentage": cpu,
               "node_memory_working_set_percentage": mem,
               "kube_node_status_allocatable_cpu_cores": 28.0}.get(nm, 10.0)
        data = [{"timeStamp": (base + dt.timedelta(hours=h)).strftime("%Y-%m-%dT%H:%M:%SZ"),
                 "average": val + (h % 5), "maximum": val + 10 + (h % 5)}
                for h in range(72)]
        out.append({"name": {"value": nm}, "timeseries": [{"data": data}]})
    return {"value": out}


DEF_POD_SEC = "/providers/Microsoft.Authorization/policySetDefinitions/a8640138-9b0a-4a28-b8cb-1666c838647d"
DEF_AUDIT_TLS = "/providers/Microsoft.Authorization/policyDefinitions/abcd1234-0000-0000-0000-000000000001"

ASSIGNMENTS = {
    S1: [{"id": "/providers/Microsoft.Management/managementGroups/corp/providers/"
                "Microsoft.Authorization/policyAssignments/k8s-baseline",
          "name": "k8s-baseline",
          "properties": {"displayName": "K8s pod security baseline",
                         "scope": "/providers/Microsoft.Management/managementGroups/corp",
                         "enforcementMode": "Default",
                         "policyDefinitionId": DEF_POD_SEC}},
         {"id": "/subscriptions/%s/providers/Microsoft.Authorization/policyAssignments/tls" % S1,
          "name": "tls",
          "properties": {"displayName": "Audit AKS HTTPS ingress",
                         "scope": "/subscriptions/%s" % S1,
                         "enforcementMode": "Default",
                         "policyDefinitionId": DEF_AUDIT_TLS}}],
    S2: [{"id": "/providers/Microsoft.Management/managementGroups/corp/providers/"
                "Microsoft.Authorization/policyAssignments/k8s-baseline",
          "name": "k8s-baseline",
          "properties": {"displayName": "K8s pod security baseline",
                         "scope": "/providers/Microsoft.Management/managementGroups/corp",
                         "enforcementMode": "Default",
                         "policyDefinitionId": DEF_POD_SEC}}],
}
DEFS = {
    DEF_POD_SEC.lower(): {"properties": {"displayName": "Kubernetes cluster pod security baseline",
                                         "metadata": {"category": "Kubernetes"}}},
    DEF_AUDIT_TLS.lower(): {"properties": {"displayName": "Audit HTTPS ingress in AKS",
                                           "metadata": {"category": "Kubernetes"}}},
}
STATES = {
    S1: [{"resourceId": CL1, "policyAssignmentId": ASSIGNMENTS[S1][1]["id"],
          "policyDefinitionId": DEF_AUDIT_TLS, "complianceState": "NonCompliant",
          "policyDefinitionAction": "audit", "policyDefinitionReferenceId": ""},
         {"resourceId": CL2, "policyAssignmentId": ASSIGNMENTS[S1][1]["id"],
          "policyDefinitionId": DEF_AUDIT_TLS, "complianceState": "Compliant",
          "policyDefinitionAction": "audit", "policyDefinitionReferenceId": ""}],
    S2: [{"resourceId": CL3, "policyAssignmentId": ASSIGNMENTS[S2][0]["id"],
          "policyDefinitionId": DEF_POD_SEC, "complianceState": "NonCompliant",
          "policyDefinitionAction": "deny", "policyDefinitionReferenceId": "podsec-1"}],
}

VERSIONS = {"values": [
    {"version": "1.30", "isDefault": False,
     "capabilities": {"supportPlan": ["AKSLongTermSupport"]},
     "patchVersions": {"1.30.9": {}}},
    {"version": "1.31", "isDefault": False,
     "capabilities": {"supportPlan": ["KubernetesOfficial", "AKSLongTermSupport"]},
     "patchVersions": {"1.31.5": {}, "1.31.6": {}}},
    {"version": "1.32", "isDefault": True,
     "capabilities": {"supportPlan": ["KubernetesOfficial", "AKSLongTermSupport"]},
     "patchVersions": {"1.32.1": {}, "1.32.2": {}}},
    {"version": "1.33", "isPreview": True,
     "capabilities": {"supportPlan": ["KubernetesOfficial"]},
     "patchVersions": {"1.33.0": {}}},
]}

ACTIVITY = {"value": [
    {"eventTimestamp": "2026-05-20T10:00:00Z",
     "operationName": {"value": "Microsoft.ContainerService/managedClusters/agentPools/write"},
     "caller": "ops@contoso.com", "status": {"value": "Succeeded"},
     "resourceId": CL1 + "/agentPools/wrk"},
    {"eventTimestamp": "2026-04-02T09:00:00Z",
     "operationName": {"value": "Microsoft.ContainerService/managedClusters/write"},
     "caller": "deploy-sp", "status": {"value": "Succeeded"}, "resourceId": CL1},
]}


def fake_request(self, method, url, *, params=None, payload=None, ok404=False,
                 min_interval=None):
    self.last_headers = {}
    low = url.lower()
    if "microsoft.resourcegraph/resources" in low:
        q = payload["query"]
        subs_l = [s.lower() for s in payload["subscriptions"]]
        if _is_rearch(q):
            data = _rearch_arg(q, subs_l)
            return {"data": data or [], "$skipToken": None}
        if "tolower(type)" in q.lower():
            data = RESOURCES
            m = re.search(r"tolower\(resourcegroup\)\s+in\s+\(([^)]*)\)", q, re.I)
            if m:
                rgs = {x.strip().strip("'").replace("''", "'").lower()
                       for x in m.group(1).split(",")}
                data = [d for d in data if d["resourceGroup"].lower() in rgs]
        elif "managedclusters" in q:
            data = CLUSTERS
        elif "microsoft.network/virtualnetworks" in q:
            data = SUBNETS
        elif "subscriptions/resourcegroups" in q:
            data = RGS
        else:
            data = SUBNAMES
        subs = [s.lower() for s in payload["subscriptions"]]
        data = [d for d in data if d["subscriptionId"].lower() in subs]
        return {"data": data, "$skipToken": None}
    if "microsoft.costmanagement/query" in low:
        return _cost_response(url, payload)
    if "/providers/microsoft.insights/metrics" in low:
        return _metrics_response(url, params)
    if "eventtypes/management/values" in low:
        return ACTIVITY
    if "kubernetesversions" in low:
        return VERSIONS
    if "policyassignments" in low:
        sub = low.split("/subscriptions/")[1].split("/")[0]
        return {"value": ASSIGNMENTS.get(sub, [])}
    if "policystates" in low:
        sub = low.split("/subscriptions/")[1].split("/")[0]
        return {"value": STATES.get(sub, [])}
    if "policysetdefinitions" in low or "policydefinitions" in low:
        for k, v in DEFS.items():
            if k in low:
                return v
        return {"properties": {"displayName": url.split("/")[-1].split("?")[0],
                               "metadata": {"category": "Other"}}}
    raise AssertionError("Unmocked URL: %s %s" % (method, url))


def fake_retail_get(url, params):
    return {"Items": [
        {"meterName": "D4s v3", "productName": "Virtual Machines Dsv3 Series",
         "unitPrice": 0.192, "type": "Consumption"},
        {"meterName": "D4s v3 Spot", "productName": "Virtual Machines Dsv3 Series",
         "unitPrice": 0.041, "type": "Consumption"},
        {"meterName": "D4s v3", "productName": "Virtual Machines Dsv3 Series Windows",
         "unitPrice": 0.38, "type": "Consumption"},
    ], "NextPageLink": None}


class FakeSession(hc.AzureSession):
    def __init__(self):
        self.last_headers = {}

    def get(self, url, **kw):
        return fake_request(self, "GET", url, **kw)

    def post(self, url, **kw):
        return fake_request(self, "POST", url, **kw)

    def put(self, url, **kw):
        return fake_request(self, "PUT", url, **kw)

    def delete(self, url, **kw):
        return fake_request(self, "DELETE", url, **kw)

    def get_paged(self, url, params=None, value_key="value"):
        return (self.get(url, params=params) or {}).get(value_key) or []

    def post_paged(self, url, payload, value_key="value"):
        return (self.post(url, payload=payload) or {}).get(value_key) or []


def fake_connect(min_interval=0.0):
    return FakeSession()


def main():
    hc.AzureSession.request = fake_request
    hc.connect = fake_connect
    armextras._retail_get = fake_retail_get
    _orig = costmgmt.CostClient.__init__
    costmgmt.CostClient.__init__ = lambda self, session, min_interval=0.0: _orig(
        self, session, 0.0)

    tmp = tempfile.mkdtemp(prefix="aksrep_")
    csv_path = os.path.join(tmp, "subscriptions.csv")
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("subscription_id,subscription_name,include\n")
        f.write("%s,contoso-platform,Y\n%s,contoso-shared,Y\n" % (S1, S2))
    out = os.path.join(tmp, "reports")

    import architecture_design
    import cluster_360
    import cluster_deepdive
    import aks_report
    import conformance
    import fleet_cost
    import fleet_inventory
    import governance
    import network_ip_capacity
    import optimization_report
    import policy_report
    import spot_cluster_report
    import subscription_rearch
    import tag_chargeback
    import utilization_idle
    import version_eol
    from azrep import sandbox
    # the sandbox helper modules must import without azure.identity / live Azure
    import azrep.kubectl  # noqa: F401
    import azrep.sandbox_clone  # noqa: F401
    import azrep.sandbox_impact  # noqa: F401
    import azrep.sandbox_k8s  # noqa: F401
    import azrep.sandbox_spot  # noqa: F401
    import azrep.sandbox_upgrade  # noqa: F401
    for mod in (architecture_design, cluster_360, cluster_deepdive, conformance,
                fleet_cost, fleet_inventory, governance,
                network_ip_capacity, optimization_report, policy_report,
                spot_cluster_report, subscription_rearch, tag_chargeback,
                utilization_idle, version_eol):
        mod.connect = fake_connect

    from openpyxl import Workbook, load_workbook

    def run(mod, argv, expect_sheets, checks=None, expect_companions=()):
        name = mod.__name__
        print("\n=== %s %s ===" % (name, " ".join(argv)))
        before = set(os.listdir(out)) if os.path.isdir(out) else set()
        mod.main(argv)
        new = [f for f in os.listdir(out) if f not in before]
        new_xlsx = [f for f in new if f.endswith(".xlsx")]
        assert len(new_xlsx) == 1, "%s: expected 1 new xlsx, got %s" % (name, new)
        for ext in expect_companions:
            assert any(f.endswith(ext) for f in new), \
                "%s: expected a companion %s file, got %s" % (name, ext, new)
        extra = [f for f in new if not f.endswith(".xlsx")
                 and not any(f.endswith(e) for e in expect_companions)]
        assert not extra, "%s: unexpected extra output files %s" % (name, extra)
        path = os.path.join(out, new_xlsx[0])
        wb = load_workbook(path)
        for sheet in expect_sheets:
            assert sheet in wb.sheetnames, "%s: missing sheet %s (has %s)" % (
                name, sheet, wb.sheetnames)
            assert wb[sheet].max_row >= 2, "%s: sheet %s is empty" % (name, sheet)
        for fn in (checks or []):
            fn(wb)
        print("    OK -> %s  sheets=%s" % (os.path.basename(path), wb.sheetnames))

    base = ["--csv", csv_path, "--out", out]

    import azrep.subs as sub_helpers
    _expect(sub_helpers.infer_env_from_name("aks-d-01") == "dev",
            "short code -d- should infer dev")
    _expect(sub_helpers.infer_env_from_name("aks-s01") == "sit",
            "short code -s01 should infer sit")
    _expect(sub_helpers.infer_env_from_name("aks-r-01") == "dr",
            "short code -r- should infer dr")

    run(fleet_inventory, base + ["--all"],
        ["ReadMe", "Clusters", "NodePools", "NetworkSecurity", "Addons", "Tags", "Summary"],
        [lambda wb: _expect(wb["Clusters"].max_row == 4, "3 clusters expected")])

    run(fleet_inventory, base + ["--env", "dev", "--cluster-prefix", "aks-dev"],
        ["ReadMe", "Clusters", "NodePools", "NetworkSecurity", "Addons", "Tags", "Summary"],
        [lambda wb: _expect(wb["Clusters"].max_row == 3,
                            "2 dev clusters expected after prefix filter")])

    run(aks_report, ["inventory"] + base + ["--env", "dev", "--cluster-prefix", "aks-dev"],
        ["ReadMe", "Clusters", "NodePools", "NetworkSecurity", "Addons", "Tags", "Summary"],
        [lambda wb: _expect(wb["Clusters"].max_row == 3,
                            "launcher should route inventory report with filters")])

    run(architecture_design, base + ["--all", "--cluster", "aks-dev-01", "--no-doc"],
        ["ReadMe", "Summary", "Clusters", "NodePools", "Network", "Subnets",
         "Resources", "ResourceCounts", "Components", "Relationships", "Diagrams"],
        [lambda wb: _expect(wb["Components"].max_row > 3,
                            "design report should include cluster components"),
         lambda wb: _expect(wb["Relationships"].max_row > 5,
                            "design report should map cluster relationships")])

    before_doc = set(os.listdir(out))
    run(architecture_design, base + ["--all", "--cluster", "aks-dev-01"],
        ["ReadMe", "Summary", "Relationships", "Diagrams"],
        expect_companions=(".md", ".drawio", ".html"))
    import xml.etree.ElementTree as ET
    new_doc = [f for f in os.listdir(out) if f not in before_doc]
    md_path = os.path.join(out, next(f for f in new_doc if f.endswith(".md")))
    with open(md_path, encoding="utf-8") as f:
        md = f.read()
    _expect("## Relationship overview" in md and "```mermaid" in md,
            "design doc should contain the Mermaid relationship overview")
    drawio_path = os.path.join(out, next(f for f in new_doc if f.endswith(".drawio")))
    tree = ET.parse(drawio_path)
    page_names = [d.get("name") for d in tree.getroot().findall("diagram")]
    _expect(page_names[0] == "Fleet relationships" and "aks-dev-01" in page_names,
            "drawio file should have an overview page and a cluster page: %s" % page_names)
    cells = tree.getroot().findall(".//mxCell")
    labels = " | ".join(c.get("value") or "" for c in cells)
    _expect("AKS: aks-dev-01" in labels and "Subnet:" in labels,
            "drawio diagram should contain cluster and subnet nodes")
    _expect(any(c.get("edge") == "1" for c in cells),
            "drawio diagram should contain relationship edges")
    html_path = os.path.join(out, next(f for f in new_doc if f.endswith(".html")))
    with open(html_path, encoding="utf-8") as f:
        html_doc = f.read()
    _expect("Fleet relationships" in html_doc and "Cluster: aks-dev-01" in html_doc,
            "HTML design view should have a fleet overview and a cluster section")
    _expect("AKS: aks-dev-01" in html_doc and 'class="card pool' in html_doc,
            "HTML design view should render cluster and node-pool cards")
    _expect("<script" not in html_doc.lower(),
            "HTML design view must stay JavaScript-free (works offline)")

    def chk_sku(wb):
        vals = [wb["SKUChanges"].cell(row=r, column=3).value
                for r in range(2, wb["SKUChanges"].max_row + 1)]
        _expect("NEW" in vals and "REMOVED" in vals,
                "SKUChanges should flag NEW and REMOVED, got %s" % vals)

    run(cluster_deepdive, base + ["--all", "--cluster", "aks-dev-01"],
        ["ReadMe", "Summary", "DailyCost", "CostByMeter", "CostByNodePool",
         "AmortizedVsActual", "SKUChanges", "NodePools", "Utilization", "ActivityLog"],
        [chk_sku])

    def chk_sections(wb):
        names = wb.sheetnames
        _expect(names[0] == "ReadMe" and names[1] == "SummaryBySubscription"
                and names[-1] == "RawMonthly",
                "sections should order ReadMe, summary, detail, reference: %s" % names)
        _expect(wb["SummaryBySubscription"].sheet_properties.tabColor is not None
                and wb["RawMonthly"].sheet_properties.tabColor is not None,
                "summary and reference tabs should be colored")
        readme = [wb["ReadMe"].cell(row=r, column=1).value
                  for r in range(1, wb["ReadMe"].max_row + 1)]
        _expect(any(str(v).startswith("Tab sections:") for v in readme),
                "ReadMe should list the tab index")

    run(fleet_cost, base + ["--all", "--actual"],
        ["ReadMe", "ClusterCosts", "PricingModelSplit", "TopMovers", "MeterChanges",
         "SummaryBySubscription", "RawMonthly"],
        [lambda wb: _expect(wb["ClusterCosts"].max_row == 4, "3 clusters in ClusterCosts"),
         chk_sections])

    def chk_eol(wb):
        ws = wb["VersionStatus"]
        col = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)].index("status") + 1
        vals = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect("OUT OF SUPPORT" in vals, "aks-dev-01 (1.29) should be OUT OF SUPPORT: %s" % vals)

    run(version_eol, base + ["--all"],
        ["ReadMe", "VersionStatus", "NodeImageAge", "SupportedVersions", "Summary"],
        [chk_eol])

    def chk_spot_detail(wb):
        _expect(wb["SpotNodePools"].max_row >= 3,
                "spot detail should include existing spot pools")
        ws = wb["SpotAssessment"]
        headers = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        check_col = headers.index("check") + 1
        checks = {ws.cell(row=r, column=check_col).value for r in range(2, ws.max_row + 1)}
        _expect("system_on_demand_pool" in checks and "spot_multi_vm_family" in checks,
                "spot assessment checks missing: %s" % checks)
        ws = wb["Candidates"]
        headers = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        _expect("Est monthly saving" in headers and ws.max_row >= 2,
                "merged spot report should screen candidates with retail savings")

    run(spot_cluster_report, base + ["--all"],
        ["ReadMe", "Summary", "SpotNodePools", "OnDemandNodePools",
         "NodePoolSkuSummary", "AutoscalerConfig", "SpotAssessment", "Candidates",
         "CostByCluster", "CostTrend", "CostByNodePool", "OtherCostItems",
         "CostByMeter", "PriceReference", "RawResourceCost"],
        [chk_spot_detail])

    run(aks_report, ["spot"] + base + ["--all", "--only-spot-clusters", "--no-retail-prices"],
        ["ReadMe", "Summary", "SpotNodePools", "SpotAssessment", "CostByCluster"],
        [lambda wb: _expect(wb["SpotNodePools"].max_row >= 3,
                            "spot alias should route to the merged spot report")])

    run(utilization_idle, base + ["--all", "--days", "3"],
        ["ReadMe", "Utilization", "IdleCandidates", "Stopped", "Summary"])

    def chk_360(wb):
        ws = wb["Cluster360"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        cat_col = hdr.index("category") + 1
        cats = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=cat_col).value
                for r in range(2, ws.max_row + 1)}
        _expect(cats.get("aks-dev-01") == "UPGRADE NOW",
                "aks-dev-01 (1.29, unsupported) should be UPGRADE NOW: %s" % cats)
        _expect(cats.get("aks-dev-02") == "STOPPED BILLING",
                "aks-dev-02 (stopped) should be STOPPED BILLING: %s" % cats)
        _expect(cats.get("aks-prod-01") == "COST HOTSPOT",
                "aks-prod-01 (+46%% MoM) should be COST HOTSPOT: %s" % cats)
        ws = wb["ActionItems"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        sev_col = hdr.index("severity") + 1
        sevs = {ws.cell(row=r, column=sev_col).value for r in range(2, ws.max_row + 1)}
        _expect("HIGH" in sevs and "MEDIUM" in sevs,
                "ActionItems should carry HIGH and MEDIUM findings: %s" % sevs)

    run(cluster_360, base + ["--all", "--days", "3"],
        ["ReadMe", "Summary", "SummaryBySubscription", "SummaryByEnvironment",
         "Cluster360", "ActionItems", "NodePools", "PricingModelSplit",
         "Utilization", "RawMonthlyCost", "CategoryLegend"],
        [chk_360])

    run(aks_report, ["360"] + base + ["--all", "--no-cost", "--no-metrics", "--days", "3"],
        ["ReadMe", "Summary", "Cluster360", "ActionItems", "NodePools",
         "CategoryLegend"],
        [lambda wb: _expect("PricingModelSplit" not in wb.sheetnames
                            and "Utilization" not in wb.sheetnames,
                            "--no-cost/--no-metrics should drop cost and metrics tabs")])

    def chk_gov(wb):
        ws = wb["Scorecard"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("not_kubenet") + 1
        vals = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=col).value
                for r in range(2, ws.max_row + 1)}
        _expect(vals.get("aks-dev-01") == "FAIL", "kubenet check should FAIL on aks-dev-01: %s" % vals)

    run(governance, base + ["--all"],
        ["ReadMe", "Scorecard", "FailDetails", "FailuresByCheck", "CheckLegend"],
        [chk_gov])

    def chk_rearch(wb):
        ws = wb["Findings"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        sev_col = hdr.index("severity") + 1
        chk_col = hdr.index("check") + 1
        sav_col = hdr.index("Est saving (USD)") + 1
        rows = [(ws.cell(row=r, column=sev_col).value,
                 ws.cell(row=r, column=chk_col).value,
                 ws.cell(row=r, column=sav_col).value)
                for r in range(2, ws.max_row + 1)]
        _expect(any(sev == "FAIL" and chk == "orphan_disks" for sev, chk, _ in rows),
                "rearch should flag the unattached disk as a FAIL: %s" % rows)
        total = wb["Summary"]
        labels = {total.cell(row=r, column=1).value: total.cell(row=r, column=2).value
                  for r in range(2, total.max_row + 1)}
        _expect(float(labels.get("Total estimated monthly savings (USD)") or 0) > 0,
                "rearch total estimated savings should be > 0: %s" % labels)

    def chk_rearch_md(md):
        _expect("Executive summary" in md, "rearch md needs an Executive summary")
        _expect("### ORPHANED" in md, "rearch md needs the ORPHANED findings section")

    before_re = set(os.listdir(out))
    run(subscription_rearch, base + ["--subs", "contoso-platform"],
        ["ReadMe", "Summary", "TopFindings", "CostTrend", "CostByRG", "Findings",
         "Orphaned", "Compute", "Storage", "StorageDisks", "PaaS&Network",
         "Advisor", "RawResources", "RawCosts"],
        [chk_rearch], expect_companions=(".md",))
    new_re = [f for f in os.listdir(out) if f not in before_re]
    re_md = os.path.join(out, next(f for f in new_re if f.endswith(".md")))
    with open(re_md, encoding="utf-8") as f:
        chk_rearch_md(f.read())

    # exactly-one-subscription guard: --all spans two subs and must exit(2)
    try:
        subscription_rearch.main(base + ["--all"])
        _expect(False, "rearch with multiple subs should exit(2)")
    except SystemExit as e:
        _expect(e.code == 2, "rearch multi-sub guard should exit(2), got %s" % e.code)

    golden_path = os.path.join(tmp, "golden.json")
    with open(golden_path, "w", encoding="utf-8") as f:
        json.dump({
            "tags": {"environment": "sandbox"},
            "cluster": {"kubernetes_version": "1.31", "private_cluster": True,
                        "disable_local_accounts": True,
                        "network": {"plugin": "azure"}},
        }, f)

    def chk_conf(wb):
        ws = wb["Scorecard"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        plug_col = hdr.index("net_plugin") + 1
        ver_col = hdr.index("version_minimum") + 1
        plug = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=plug_col).value
                for r in range(2, ws.max_row + 1)}
        ver = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=ver_col).value
               for r in range(2, ws.max_row + 1)}
        _expect(plug.get("aks-dev-01") == "FAIL" and plug.get("aks-prod-01") == "PASS",
                "kubenet cluster should drift from the azure-CNI golden: %s" % plug)
        _expect(ver.get("aks-dev-01") == "FAIL" and ver.get("aks-dev-02") == "PASS",
                "1.29 cluster should fail the 1.31 version floor: %s" % ver)

    run(conformance, base + ["--all", "--golden", golden_path],
        ["ReadMe", "Scorecard", "FailDetails", "FailuresByRule", "RuleLegend"],
        [chk_conf])

    def chk_pol(wb):
        ws = wb["KubernetesBlindSpots"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("status") + 1
        vals = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=col).value
                for r in range(2, ws.max_row + 1)}
        _expect(vals.get("aks-prod-01") == "BLIND SPOT",
                "prod cluster without policy addon should be BLIND SPOT: %s" % vals)

    run(policy_report, base + ["--all"],
        ["ReadMe", "Assignments", "ClusterCompliance", "NonCompliantDetail",
         "KubernetesBlindSpots", "Summary"],
        [chk_pol])

    def chk_network(wb):
        ws = wb["SubnetCapacity"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("status") + 1
        vals = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect("CRITICAL" in vals or "WARN" in vals,
                "network report should flag constrained subnets: %s" % vals)

    run(network_ip_capacity, base + ["--all"],
        ["ReadMe", "ClusterNetwork", "SubnetCapacity", "PoolSubnetUse", "Issues", "Summary"],
        [chk_network])

    def chk_tags(wb):
        ws = wb["MissingTags"]
        vals = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
        _expect("owner" in vals and "costcenter" in vals,
                "tag report should flag missing owner/costcenter: %s" % vals)

    run(tag_chargeback, base + ["--all"],
        ["ReadMe", "TagMatrix", "MissingTags", "TagCoverage", "TagValues", "RawTags", "Summary"],
        [chk_tags])

    def chk_opt(wb):
        ws = wb["SavingsCandidates"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("candidate") + 1
        vals = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect("STOPPED_BILLING" in vals or "SPOT_REVIEW" in vals,
                "optimization report should flag stopped/spot candidates: %s" % vals)

    run(optimization_report, base + ["--all", "--days", "3"],
        ["ReadMe", "Summary", "SavingsCandidates", "ClusterCostUtilization",
         "PricingModelSplit", "RawMonthly"],
        [chk_opt])

    prisma_path = os.path.join(tmp, "prisma.xlsx")
    vuln_wb = Workbook()
    vuln_ws = vuln_wb.active
    vuln_ws.title = "Vulnerabilities"
    vuln_ws.append(["CVE IDs", "Severity", "Package Name", "Installed Version",
                    "Fixed Version", "Package Category", "Container Image",
                    "Cluster Name", "Entity Type", "Image Layer", "OS Distro"])
    vuln_ws.append(["CVE-2026-10001", "High", "openssl", "3.0.1", "3.0.8",
                    "OS Package", "registry.example.com/app:1.0",
                    "aks-dev-01", "image", "/usr/lib/libssl.so", "Ubuntu"])
    vuln_ws.append(["CVE-2026-10002", "Critical", "spring-core", "5.3.0", "6.0.0",
                    "Maven", "registry.example.com/app:1.0",
                    "aks-dev-01", "container", "/app/lib/spring-core.jar", ""])
    vuln_ws.append(["CVE-2026-10003", "High", "openjdk-17-jre", "17.0.1", "17.0.9",
                    "OS Package", "registry.example.com/app:1.0",
                    "aks-dev-01", "image", "/usr/lib/jvm/java-17-openjdk", ""])
    vuln_wb.save(prisma_path)

    def chk_vuln(wb):
        ws = wb["Classification"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        layer_col = hdr.index("layer") + 1
        layers = {ws.cell(row=r, column=layer_col).value
                  for r in range(2, ws.max_row + 1)}
        _expect({"base_image", "application", "platform"} <= layers,
                "vulnerability report should classify all layers: %s" % layers)

    run(aks_report, ["vulnerabilities", "--prisma", prisma_path,
                     "--classification-rules", os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                                            "vulnerability_classification.example.json"),
                     "--offline", "--out", out],
        ["ReadMe", "Summary", "PrismaFindings", "Classification", "Remediation",
         "ByImage", "ByPackage", "ByLayer", "CVEReference", "ClassificationRules",
         "InputColumns"],
        [chk_vuln])

    import container_os_eol
    import aks_lifecycle

    def eol_fixture(slug, timeout=30):
        day = dt.timedelta(days=1)
        return [
            {"cycle": "9", "latest": "9.9", "lts": slug == "nodejs",
             "releaseDate": str(TODAY - 400 * day), "support": str(TODAY + 200 * day),
             "eol": str(TODAY + 400 * day)},
            {"cycle": "8", "latest": "8.8", "lts": False,
             "releaseDate": str(TODAY - 700 * day), "support": str(TODAY - 30 * day),
             "eol": str(TODAY + 60 * day)},
            {"cycle": "7", "latest": "7.7", "lts": False,
             "releaseDate": str(TODAY - 900 * day), "support": str(TODAY - 300 * day),
             "eol": str(TODAY - 30 * day)},
        ]

    container_os_eol.fetch_product = eol_fixture

    def chk_eol(wb):
        ws = wb["EolRadar"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("status") + 1
        vals = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect({"SUPPORTED", "EOL <90 DAYS", "EOL"} <= vals,
                "EOL radar should mix supported/eol-soon/eol rows: %s" % vals)

    run(container_os_eol, ["--out", out],
        ["ReadMe", "Summary", "EolRadar", "OsBaseImages", "LanguageRuntimes",
         "RawLifecycle"], [chk_eol])

    ga_month = (TODAY - dt.timedelta(days=200)).strftime("%b %Y")
    eol_month = (TODAY + dt.timedelta(days=300)).strftime("%b %Y")
    gone_month = (TODAY - dt.timedelta(days=40)).strftime("%b %Y")
    versions_html = """
    <h2>AKS Kubernetes release calendar</h2>
    <table><tr><th>Kubernetes version</th><th>Upstream release</th><th>AKS preview</th>
    <th>AKS GA</th><th>End of life</th><th>Platform support</th></tr>
    <tr><td>1.33</td><td>Apr 2025</td><td>May 2025</td><td>%s</td><td>%s</td>
    <td>Until 1.37 GA</td></tr>
    <tr><td>1.31</td><td>Aug 2024</td><td>Sep 2024</td><td>Nov 2024</td><td>%s</td>
    <td>Until 1.35 GA</td></tr></table>
    <h3>LTS versions</h3>
    <table><tr><th>Kubernetes version</th><th>Upstream release</th><th>AKS preview</th>
    <th>AKS GA</th><th>End of life</th><th>LTS End of life</th></tr>
    <tr><td>1.30</td><td>Apr 2024</td><td>Jun 2024</td><td>Jul 2024</td><td>%s</td>
    <td>%s</td></tr></table>
    <h3>Kubernetes 1.33</h3>
    <table><tr><th>AKS managed add-ons (addon)</th><th>AKS components (ccp)</th>
    <th>OS components</th><th>Breaking changes from Kubernetes 1.32.0</th></tr>
    <tr><td>coredns v1.12.1</td><td>cluster-autoscaler v1.33.0</td>
    <td>Ubuntu 22.04</td><td>calico v3.30 bump</td></tr></table>
    """ % (ga_month, eol_month, gone_month, gone_month, eol_month)
    integrations_html = """
    <h2>Available add-ons</h2>
    <table><tr><th>Name</th><th>Description</th><th>Articles</th><th>GitHub</th></tr>
    <tr><td>keda</td><td>Event-driven autoscaling</td>
    <td><a href="keda-about">KEDA add-on</a></td>
    <td><a href="https://github.com/kedacore/keda">GitHub</a></td></tr></table>
    <h2>Open-source and third-party integrations</h2>
    <table><tr><th>Name</th><th>Description</th><th>More details</th></tr>
    <tr><td><a href="https://helm.sh">Helm</a></td><td>Packaging tool</td>
    <td><a href="quickstart-helm">Quickstart</a></td></tr></table>
    """
    release_body = """## Release Notes - 2026-01-02

### Announcements
* Feature X retired on January 1, 2026. Migrate to feature Z.
* The legacy flag Y has been deprecated and will be removed.
* [Feature Z](https://learn.microsoft.com/azure/aks/z) is now generally available.

### Features
* Thing A is now generally available in all regions.

### Preview features
* Thing B is now in public preview.

### Behavioral changes
* Default for C changed from off to on.

### Bug fixes
* Fixed D crashing.

### Component updates
* Bumped E to v2.0.0.
"""
    aks_lifecycle.fetch_html = lambda url, timeout=30: (
        versions_html if "supported-kubernetes-versions" in url else integrations_html)
    aks_lifecycle.fetch_releases = lambda count=30, timeout=30: [
        {"tag_name": "2026-01-02", "published_at": "2026-01-05T00:00:00Z",
         "body": release_body}]

    def chk_lifecycle(wb):
        ws = wb["Announcements"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("kind") + 1
        kinds = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect({"RETIREMENT", "DEPRECATION", "GA"} <= kinds,
                "lifecycle report should classify announcements: %s" % kinds)
        ws = wb["ReleaseCalendar"]
        hdr = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        col = hdr.index("status") + 1
        statuses = {ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)}
        _expect("GA" in statuses and "EOL" in statuses,
                "release calendar should compute GA and EOL statuses: %s" % statuses)

    run(aks_lifecycle, ["--out", out],
        ["ReadMe", "Summary", "ReleaseCalendar", "Announcements", "GAFeatures",
         "PreviewFeatures", "BehaviorChanges", "Addons", "OpenSourceIntegrations",
         "BreakingChanges", "ComponentUpdates", "RawReleaseNotes"], [chk_lifecycle])

    md_path = os.path.join(tmp, "sample.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Sample Report\n\n")
        f.write("Short paragraph for document export.\n\n")
        f.write("| Item | Value |\n|---|---|\n| Clusters | 3 |\n\n")
        f.write("- one\n- two\n")
    export_dir = os.path.join(tmp, "exports")
    aks_report.main(["convert", md_path, "--to", "all", "--out-dir", export_dir,
                     "--output-name", "sample"])
    _expect(os.path.exists(os.path.join(export_dir, "sample.docx")),
            "Markdown DOCX export missing")
    _expect(os.path.exists(os.path.join(export_dir, "sample.pdf")),
            "Markdown PDF export missing")

    sandbox_cfg = {
        "subscription_id": S1,
        "subscription_name": "contoso-sandbox",
        "environment": "sandbox",
        "resource_group": "rg-aks-sandbox-dev",
        "location": "eastus",
        "cluster": {
            "name": "aks-sbx-policy-01",
            "network": {"plugin": "azure", "plugin_mode": "overlay", "policy": "azure"},
            "node_pools": [
                {"name": "sys", "mode": "System", "vm_size": "Standard_D4s_v5",
                 "count": 1, "autoscaling": True, "min_count": 1, "max_count": 2}
            ],
        },
        "policies": {
            "assignment_scope": "resource_group",
            "definitions": [],
            "assignments": [],
        },
    }
    sandbox_path = os.path.join(tmp, "sandbox.json")
    with open(sandbox_path, "w", encoding="utf-8") as f:
        json.dump(sandbox_cfg, f)
    aks_report.main(["sandbox", "plan", sandbox_path])
    template = sandbox.build_aks_template(sandbox_cfg)
    aks_res = template["resources"][0]
    _expect(aks_res["name"] == "aks-sbx-policy-01", "sandbox template cluster name mismatch")
    _expect(aks_res["properties"]["agentPoolProfiles"][0]["mode"] == "System",
            "sandbox template needs a System pool")

    print("\nALL SMOKE TESTS PASSED  (outputs in %s)" % out)


def _expect(cond, msg):
    assert cond, msg


if __name__ == "__main__":
    main()
