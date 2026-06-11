"""Offline test for spot_split_design.py against a Korea-style fixture:
one dev cluster with team-dedicated node pools (labels, taints, namespaces).

  uv run python tests/test_spot_split.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import azrep.http_client as hc

S1 = "11111111-1111-1111-1111-111111111111"
KR = ("/subscriptions/%s/resourcegroups/rg-kr-dev/providers/"
      "microsoft.containerservice/managedclusters/aks-kr-dev-01" % S1)


def pool(name, mode, size, count, labels=None, taints=None, priority="Regular",
         zones=("1", "2"), autoscale=False, min_c=None, max_c=None, os_type="Linux"):
    p = {"name": name, "mode": mode, "vmSize": size, "count": count,
         "scaleSetPriority": priority, "availabilityZones": list(zones),
         "enableAutoScaling": autoscale, "osType": os_type, "osSKU": "Ubuntu",
         "osDiskType": "Managed", "osDiskSizeGB": 128, "maxPods": 110,
         "orchestratorVersion": "1.31.5", "currentOrchestratorVersion": "1.31.5",
         "nodeImageVersion": "AKSUbuntu-2204gen2containerd-202601.07.0",
         "type": "VirtualMachineScaleSets", "powerState": {"code": "Running"}}
    if labels:
        p["nodeLabels"] = labels
    if taints:
        p["nodeTaints"] = taints
    if autoscale:
        p["minCount"], p["maxCount"] = min_c or 1, max_c or max(count, 5)
    if priority == "Spot":
        p["spotMaxPrice"] = -1
        p["scaleSetEvictionPolicy"] = "Delete"
        p["nodeTaints"] = (taints or []) + [
            "kubernetes.azure.com/scalesetpriority=spot:NoSchedule"]
    return p


KOREA_CLUSTER = {
    "id": KR, "name": "aks-kr-dev-01", "resourceGroup": "rg-kr-dev",
    "subscriptionId": S1, "location": "koreacentral",
    "tags": {"environment": "dev"},
    "kubernetesVersion": "1.31.5", "currentKubernetesVersion": "1.31.5",
    "provisioningState": "Succeeded", "powerState": "Running",
    "skuName": "Base", "skuTier": "Standard", "supportPlan": "KubernetesOfficial",
    "nodeResourceGroup": "MC_rg-kr-dev_aks-kr-dev-01_koreacentral",
    "dnsPrefix": "akskrdev01", "fqdn": "akskrdev01.hcp.koreacentral.azmk8s.io",
    "privateFQDN": "", "enableRBAC": True, "disableLocalAccounts": True,
    "identityType": "SystemAssigned", "servicePrincipalClientId": "",
    "agentPoolProfiles": [
        pool("sys", "System", "Standard_D4s_v5", 2),
        # team via node label
        pool("paypool", "User", "Standard_D8s_v3", 4,
             labels={"team": "payments"},
             taints=["dedicated=payments:NoSchedule"],
             autoscale=True, min_c=2, max_c=6),
        # team via taint only
        pool("ordpool", "User", "Standard_D4s_v3", 3,
             taints=["dedicated=orders:NoSchedule"]),
        # team via name heuristic + teams.csv override
        pool("dataeng", "User", "Standard_D8as_v4", 2),
        # already spot - must not be split again
        pool("shspot", "User", "Standard_D4as_v4", 2, priority="Spot",
             labels={"kubernetes.azure.com/scalesetpriority": "spot"}),
        # windows pool - listed but not split
        pool("winapp", "User", "Standard_D4s_v3", 1, os_type="Windows"),
    ],
    "networkProfile": {"networkPlugin": "azure", "networkPolicy": "calico",
                       "loadBalancerSku": "standard", "outboundType": "loadBalancer"},
    "apiServerAccessProfile": {"authorizedIPRanges": ["10.0.0.0/8"]},
    "aadProfile": {"managed": True, "enableAzureRBAC": True},
    "addonProfiles": {"azurepolicy": {"enabled": True}, "omsagent": {"enabled": True}},
    "autoUpgradeProfile": {"upgradeChannel": "patch", "nodeOSUpgradeChannel": "NodeImage"},
    "securityProfile": {}, "oidcIssuerProfile": {"enabled": True},
    "autoScalerProfile": {"expander": "least-waste",
                          "balance-similar-node-groups": "false"},
}


def fake_request(self, method, url, *, params=None, payload=None, ok404=False,
                 min_interval=None):
    self.last_headers = {}
    low = url.lower()
    if "microsoft.resourcegraph/resources" in low:
        q = payload["query"]
        if "managedclusters" in q:
            data = [KOREA_CLUSTER]
        elif "subscriptions/resourcegroups" in q:
            data = [{"subscriptionId": S1, "name": "rg-kr-dev",
                     "tags": {"environment": "dev"}}]
        else:
            data = [{"subscriptionId": S1, "name": "contoso-kr"}]
        return {"data": data, "$skipToken": None}
    raise AssertionError("Unmocked URL: %s %s" % (method, url))


class FakeSession(hc.AzureSession):
    def __init__(self):
        super().__init__(credential=None)


def fake_connect(min_interval=0.0):
    return FakeSession()


def fake_retail_get(url, params):
    return {"Items": [
        {"meterName": "D8s v3", "productName": "Virtual Machines Dsv3 Series",
         "unitPrice": 0.48, "type": "Consumption"},
        {"meterName": "D8s v3 Spot", "productName": "Virtual Machines Dsv3 Series",
         "unitPrice": 0.10, "type": "Consumption"},
    ], "NextPageLink": None}


def expect(cond, msg):
    assert cond, msg


def main():
    hc.AzureSession.request = fake_request
    hc.connect = fake_connect
    import azrep.armextras as armextras
    armextras._retail_get = fake_retail_get

    import spot_split_design
    spot_split_design.connect = fake_connect

    from openpyxl import load_workbook

    tmp = tempfile.mkdtemp(prefix="spotsplit_")
    csv_path = os.path.join(tmp, "subscriptions.csv")
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("subscription_id,subscription_name,include\n")
        f.write("%s,contoso-kr,Y\n" % S1)
    teams_csv = os.path.join(tmp, "teams.csv")
    with open(teams_csv, "w", encoding="utf-8") as f:
        f.write("pool,team,namespaces,workload_type,criticality\n")
        f.write("dataeng,data-engineering,data;etl,batch,1\n")
    out = os.path.join(tmp, "reports")

    path = spot_split_design.main(["--csv", csv_path, "--out", out,
                                   "--cluster", "aks-kr-dev-01",
                                   "--teams", teams_csv])
    wb = load_workbook(path)
    for sheet in ("ReadMe", "CurrentState", "TeamMapping", "FutureStatePools",
                  "AzCommands", "WorkloadChanges", "Savings", "NotSplit",
                  "ClusterPrereqs", "RolloutPlan", "Risks"):
        expect(sheet in wb.sheetnames, "missing sheet %s (has %s)" % (sheet, wb.sheetnames))

    ws = wb["TeamMapping"]
    hdr = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}
    teams = {ws.cell(row=r, column=hdr["pool"]).value:
             ws.cell(row=r, column=hdr["team"]).value
             for r in range(2, ws.max_row + 1)}
    expect(teams["paypool"] == "payments", "label attribution failed: %s" % teams)
    expect(teams["ordpool"] == "orders", "taint attribution failed: %s" % teams)
    expect(teams["dataeng"] == "data-engineering", "teams.csv override failed: %s" % teams)

    ws = wb["FutureStatePools"]
    hdr = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}
    rows = [{k: ws.cell(row=r, column=j).value for k, j in hdr.items()}
            for r in range(2, ws.max_row + 1)]
    expect(len(rows) == 3, "expected 3 split plans (pay/ord/dataeng), got %d" % len(rows))
    pay = next(r for r in rows if r["team"] == "payments")
    expect(pay["od_keep_nodes"] == 1 and pay["spot_initial_nodes"] == 3,
           "payments split math wrong: %s" % pay)
    expect(len(str(pay["spot_pool"])) <= 12 and str(pay["spot_pool"]).endswith("sp"),
           "spot pool name invalid: %s" % pay["spot_pool"])

    ws = wb["AzCommands"]
    cmds = [str(ws.cell(row=r, column=5).value) for r in range(2, ws.max_row + 1)]
    expect(sum("az aks nodepool add" in c for c in cmds) == 3, "3 add commands expected")
    expect(any("--priority Spot" in c and "--spot-max-price -1" in c for c in cmds),
           "spot flags missing in add commands")
    expect(any("dedicated=payments:NoSchedule" in c for c in cmds),
           "team taint must be preserved on the spot pool")

    ws = wb["NotSplit"]
    skipped = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    expect({"sys", "shspot", "winapp"} <= skipped, "sys/spot/windows must be skipped: %s" % skipped)

    ws = wb["WorkloadChanges"]
    y = str(ws.cell(row=2, column=3).value)
    expect("scalesetpriority" in y and "topologySpreadConstraints" in y
           and "PodDisruptionBudget" in y, "workload YAML incomplete")

    md_path = os.path.splitext(path)[0] + ".md"
    expect(os.path.exists(md_path), "design doc .md missing")
    md = open(md_path, encoding="utf-8").read()
    expect(md.count("```mermaid") == 2, "expected current+future mermaid diagrams")
    expect("Present state" in md and "Future state" in md, "md sections missing")

    print("\nALL SPOT-SPLIT TESTS PASSED")
    print("workbook: %s" % path)
    print("design doc: %s" % md_path)


if __name__ == "__main__":
    main()
