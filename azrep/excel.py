"""Multi-tab xlsx report helpers (openpyxl). Consistent styling: Arial 10,
dark-blue header, frozen header row, autofilter, number formats, conditional
formatting and simple charts.

Every workbook follows the same four-section layout. Tag each sheet with
`section=` and save() enforces the order, colors the tabs and appends a tab
index to the ReadMe sheet:

  intro (blue)     ReadMe - what the report is, scope, generated timestamp
  summary (green)  Summary first, then SummaryBy<Dim> breakdowns
  detail           findings and per-entity tables (default section)
  reference (gray) Raw* extracts and lookup/legend tabs, always last
"""
import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = "1F4E78"
FAIL_FILL = PatternFill("solid", start_color="F8CBAD")
WARN_FILL = PatternFill("solid", start_color="FFE699")
MONEY = "#,##0.00"
MONEY0 = "#,##0"
PCT = "0.0%"
INT = "#,##0"
DATA_FONT = Font(name=FONT, size=10)
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")

SECTION_ORDER = ("intro", "summary", "detail", "reference")
SECTION_TAB_COLORS = {"intro": "1F4E78", "summary": "70AD47", "reference": "A6A6A6"}
SECTION_LABELS = {"summary": "Summary", "detail": "Detail", "reference": "Reference"}


def new_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    wb._azrep_sections = {}
    return wb


def _set_section(wb, ws, section):
    if section not in SECTION_ORDER:
        raise ValueError("unknown section %r (expected one of %s)" % (section, SECTION_ORDER))
    getattr(wb, "_azrep_sections", {})[ws.title] = section
    color = SECTION_TAB_COLORS.get(section)
    if color:
        ws.sheet_properties.tabColor = color


def add_readme(wb, title, lines):
    ws = wb.create_sheet("ReadMe", 0)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    r = 3
    for line in lines:
        ws.cell(row=r, column=1, value=line).font = DATA_FONT
        r += 1
    ws.column_dimensions["A"].width = 150
    _set_section(wb, ws, "intro")
    return ws


def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, (dict, list, set, tuple)):
        return str(v)
    if pd.isna(v):
        return None
    return v


def add_table(wb, name, df, money_cols=(), pct_cols=(), int_cols=(), formats=None,
              fail_cols=(), warn_values=("WARN",), fail_values=("FAIL", "OUT OF SUPPORT",
              "BLIND SPOT", "NonCompliant", "STOPPED", "IDLE"), colorscale_cols=(),
              max_width=55, index=False, section="detail"):
    """Write a DataFrame as a styled sheet. Cells whose string starts with '='
    are written as Excel formulas. `section` places the sheet within the
    standard workbook layout (see module docstring); save() enforces order."""
    ws = wb.create_sheet(str(name)[:31])
    _set_section(wb, ws, section)
    if index:
        df = df.reset_index()
    cols = [str(c) for c in df.columns]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = HDR_FONT
        cell.fill = PatternFill("solid", start_color=HDR_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False, name=None), 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=_clean(v)).font = DATA_FONT
    nrows = len(df) + 1

    fmt = {}
    for c in money_cols:
        fmt[c] = MONEY
    for c in pct_cols:
        fmt[c] = PCT
    for c in int_cols:
        fmt[c] = INT
    fmt.update(formats or {})
    for c, f in fmt.items():
        if c in cols:
            L = get_column_letter(cols.index(c) + 1)
            for r in range(2, nrows + 1):
                ws["%s%d" % (L, r)].number_format = f

    for j, c in enumerate(cols, 1):
        try:
            longest = max([len(str(c))] + [len(str(_clean(v) or "")) for v in df[df.columns[j - 1]].head(200)])
        except Exception:
            longest = len(str(c))
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, longest + 2), max_width)

    ws.freeze_panes = "A2"
    if nrows > 1:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), nrows)

    if nrows > 1:
        for c in fail_cols:
            if c in cols:
                L = get_column_letter(cols.index(c) + 1)
                rng = "%s2:%s%d" % (L, L, nrows)
                for v in fail_values:
                    ws.conditional_formatting.add(rng, CellIsRule(
                        operator="equal", formula=['"%s"' % v], fill=FAIL_FILL))
                for v in warn_values:
                    ws.conditional_formatting.add(rng, CellIsRule(
                        operator="equal", formula=['"%s"' % v], fill=WARN_FILL))
        for c in colorscale_cols:
            if c in cols:
                L = get_column_letter(cols.index(c) + 1)
                ws.conditional_formatting.add(
                    "%s2:%s%d" % (L, L, nrows),
                    ColorScaleRule(start_type="min", start_color="63BE7B",
                                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                   end_type="max", end_color="F8696B"))
    return ws


def add_total_row(ws, df, sum_cols, label_col=None):
    """Append a bold Total row with =SUM() formulas (keeps the workbook dynamic)."""
    cols = [str(c) for c in df.columns]
    r = len(df) + 2
    lab = 1 if label_col is None else cols.index(label_col) + 1
    cell = ws.cell(row=r, column=lab, value="Total")
    cell.font = Font(name=FONT, size=10, bold=True)
    for c in sum_cols:
        if c in cols:
            j = cols.index(c) + 1
            L = get_column_letter(j)
            f = ws.cell(row=r, column=j, value="=SUM(%s2:%s%d)" % (L, L, r - 1))
            f.font = Font(name=FONT, size=10, bold=True)
            f.number_format = ws["%s2" % L].number_format or MONEY
    return r


def add_line_chart(ws, title, nrows, first_data_col, last_data_col, anchor,
                   y_title="USD", cat_col=1):
    ch = LineChart()
    ch.title = title
    ch.height, ch.width = 9, 24
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    data = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=nrows)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)
    return ch


def add_bar_chart(ws, title, nrows, data_col, anchor, y_title="USD", cat_col=1):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.height, ch.width = 9, 24
    ch.y_axis.title = y_title
    data = Reference(ws, min_col=data_col, max_col=data_col, min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=nrows)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)
    return ch


def add_grouped_bar_chart(ws, title, nrows, first_data_col, last_data_col, anchor,
                          y_title="USD", cat_col=1):
    """Clustered (side-by-side) column chart over a contiguous range of data
    columns - one series per column, grouped per category. Mirrors
    add_line_chart's data-range signature; use for before/after comparisons."""
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "clustered"
    ch.title = title
    ch.height, ch.width = 9, 24
    ch.y_axis.title = y_title
    data = Reference(ws, min_col=first_data_col, max_col=last_data_col,
                     min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=nrows)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)
    return ch


SCORECARD_FILL = {"good": "C6EFCE", "warn": "FFEB9C", "bad": "FFC7CE",
                  "neutral": "DDEBF7"}


def add_scorecard(wb, name, cards, section="summary", per_row=3, title=None):
    """Render KPI cards (label / big value / caption) as merged blocks for an
    exec-readable one-pager, grid-wrapped `per_row` across. `cards` is a list of
    dicts with keys label, value, caption, rag (rag in good/warn/bad/neutral ->
    soft fill, black text). Returns the worksheet; cards span 4 cols x 4 rows each
    with a 1-cell gutter. Used instead of a tall data table when the audience is
    non-technical (see spot_savings Scorecard)."""
    ws = wb.create_sheet(str(name)[:31])
    _set_section(wb, ws, section)
    start_row, start_col, cw, ch, gap = 2, 2, 4, 4, 1
    if title:
        ws.cell(row=1, column=start_col, value=title).font = Font(
            name=FONT, size=12, bold=True)
    label_font = Font(name=FONT, size=9, bold=True, color="404040")
    value_font = Font(name=FONT, size=20, bold=True, color="000000")
    cap_font = Font(name=FONT, size=8, color="595959")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    last_row = start_row
    for i, card in enumerate(cards):
        r0 = start_row + (i // per_row) * (ch + gap)
        c0 = start_col + (i % per_row) * (cw + gap)
        fill = PatternFill("solid", start_color=SCORECARD_FILL.get(
            str(card.get("rag", "neutral")), SCORECARD_FILL["neutral"]))
        for rr in range(r0, r0 + ch):
            for cc in range(c0, c0 + cw):
                ws.cell(row=rr, column=cc).fill = fill
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + cw - 1)
        lc = ws.cell(row=r0, column=c0, value=card.get("label", ""))
        lc.font, lc.alignment = label_font, center
        ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 2, end_column=c0 + cw - 1)
        vc = ws.cell(row=r0 + 1, column=c0, value=card.get("value", ""))
        vc.font, vc.alignment = value_font, center
        ws.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c0 + cw - 1)
        pc = ws.cell(row=r0 + 3, column=c0, value=card.get("caption", ""))
        pc.font, pc.alignment = cap_font, center
        last_row = max(last_row, r0 + ch)
    for cidx in range(start_col, start_col + per_row * (cw + gap)):
        ws.column_dimensions[get_column_letter(cidx)].width = 13
    ws._azrep_scorecard_next_row = last_row + 1
    return ws


def _finalize_sections(wb):
    sections = getattr(wb, "_azrep_sections", None)
    if not sections:
        return
    rank = {s: i for i, s in enumerate(SECTION_ORDER)}
    wb._sheets.sort(key=lambda ws: rank[sections.get(ws.title, "detail")])
    if "ReadMe" not in wb.sheetnames:
        return
    ws = wb["ReadMe"]
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="Tab sections:").font = Font(
        name=FONT, size=10, bold=True)
    for section in SECTION_ORDER[1:]:
        names = [s.title for s in wb._sheets
                 if sections.get(s.title, "detail") == section]
        if names:
            r += 1
            ws.cell(row=r, column=1, value="  %-10s %s" % (
                SECTION_LABELS[section] + ":", ", ".join(names))).font = DATA_FONT


def save(wb, path):
    _finalize_sections(wb)
    wb.save(path)
    return path
